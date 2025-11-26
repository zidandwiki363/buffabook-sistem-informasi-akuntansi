import streamlit as st
import openpyxl
from openpyxl import Workbook, load_workbook
import pandas as pd
import os
from datetime import datetime
import sqlite3
import hashlib
import re

def init_auth_db():
    """Initialize SQLite database for authentication"""
    conn = sqlite3.connect('buffabook_auth.db')
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

def hash_password(password):
    """Hash password using SHA-256"""
    return hashlib.sha256(password.encode()).hexdigest()

def create_user(email, password):
    """Create new user in database"""
    try:
        conn = sqlite3.connect('buffabook_auth.db')
        c = conn.cursor()
        password_hash = hash_password(password)
        c.execute('INSERT INTO users (email, password_hash) VALUES (?, ?)', 
                 (email, password_hash))
        conn.commit()
        conn.close()
        return True
    except sqlite3.IntegrityError:
        return False  
    except Exception as e:
        st.error(f"Error creating user: {e}")
        return False

def verify_user(email, password):
    """Verify user credentials"""
    try:
        conn = sqlite3.connect('buffabook_auth.db')
        c = conn.cursor()
        password_hash = hash_password(password)
        c.execute('SELECT * FROM users WHERE email = ? AND password_hash = ?', 
                 (email, password_hash))
        user = c.fetchone()
        conn.close()
        return user is not None
    except Exception as e:
        st.error(f"Error verifying user: {e}")
        return False

def is_valid_email(email):
    """Validate email format"""
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None

init_auth_db()

st.set_page_config(
    page_title="BuffaBook - Peternakan Kerbau",
    page_icon="🐃",
    layout="wide",
    initial_sidebar_state="expanded"
)

if 'logged_in' not in st.session_state:
    st.session_state.logged_in = False
if 'user_email' not in st.session_state:
    st.session_state.user_email = None
if 'show_login' not in st.session_state:
    st.session_state.show_login = True  

if not st.session_state.logged_in:
    st.markdown("""
    <style>
        /* Set background color on main Streamlit app container for auth page */
        [data-testid="stAppViewContainer"] {
            background-color: #FDF3B9 !important;
        }
        .auth-container {
            background: #FDF3B9 ;
            padding: 2rem;
            border-radius: 10px;
            box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
            margin: 2rem auto;
            max-width: 400px;
        }
        .auth-header {
            text-align: center;
            color: #80644B;
            margin-bottom: 2rem;
        }
        .stButton>button {
            width: 100%;
            background-color: #8B6F47 ;
            color: white ;
            border: none ;
        }
        .stButton>button:hover {
            background-color: #5A4634 ;
        }
        .switch-auth {
            text-align: center;
            margin-top: 1rem;
        }
        .switch-auth button {
            background: none ;
            border: none;
            color: #8B6F47;
            text-decoration: underline;
            cursor: pointer;
        }
    </style>
    """, unsafe_allow_html=True)
    
    st.markdown('<div class="auth-header"><h1>🐃 BuffaBook</h1><p>Sistem Pencatatan Peternakan Kerbau</p></div>', unsafe_allow_html=True)
    
    if st.session_state.show_login:
        st.subheader("🔐 Login")
        
        with st.form("login_form"):
            email = st.text_input("Email", placeholder="nama@email.com")
            password = st.text_input("Password", type="password", placeholder="Minimal 6 karakter")
            login_submit = st.form_submit_button("Login")
            
            if login_submit:
                if not email or not password:
                    st.error("Email dan password harus diisi!")
                elif not is_valid_email(email):
                    st.error("Format email tidak valid!")
                elif len(password) < 6:
                    st.error("Password minimal 6 karakter!")
                else:
                    if verify_user(email, password):
                        st.session_state.logged_in = True
                        st.session_state.user_email = email
                        st.success("Login berhasil!")
                        st.rerun()
                    else:
                        st.error("Email atau password salah!")
        
        st.markdown('<div class="switch-auth">', unsafe_allow_html=True)
        if st.button("Belum punya akun? Daftar di sini"):
            st.session_state.show_login = False
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
    
    else:
        st.subheader("📝 Register")
        
        with st.form("register_form"):
            email = st.text_input("Email", placeholder="nama@email.com")
            password = st.text_input("Password", type="password", placeholder="Minimal 6 karakter")
            confirm_password = st.text_input("Konfirmasi Password", type="password", placeholder="Ulangi password")
            register_submit = st.form_submit_button("Daftar")
            
            if register_submit:
                if not email or not password or not confirm_password:
                    st.error("Semua field harus diisi!")
                elif not is_valid_email(email):
                    st.error("Format email tidak valid!")
                elif len(password) < 6:
                    st.error("Password minimal 6 karakter!")
                elif password != confirm_password:
                    st.error("Password dan konfirmasi password tidak sama!")
                else:
                    if create_user(email, password):
                        st.success("Pendaftaran berhasil! Silakan login.")
                        st.session_state.show_login = True
                        st.rerun()
                    else:
                        st.error("Email sudah terdaftar! Silakan gunakan email lain.")
        
        st.markdown('<div class="switch-auth">', unsafe_allow_html=True)
        if st.button("Sudah punya akun? Login di sini"):
            st.session_state.show_login = True
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
    
    st.markdown('</div>', unsafe_allow_html=True)
    st.stop()

ACCOUNTS = {
    "1-10000": "Kas",
    "1-11000": "Piutang",
    "1-12000": "Persediaan Kerbau Dewasa Jantan",
    "1-12100": "Persediaan Kerbau Dewasa Betina",
    "1-12200": "Persediaan Kerbau Remaja Jantan",
    "1-12300": "Persediaan Kerbau Remaja Betina",
    "1-12400": "Persediaan Anak Kerbau Jantan",
    "1-12500": "Persediaan Anak Kerbau Betina",
    "1-20000": "Kendaraan",
    "1-21000": "Kandang",
    "1-22000": "Peralatan",
    "1-23000": "Akumulasi Penyusutan Kendaraan",
    "1-23100": "Akumulasi Penyusutan Kandang",
    "1-23200": "Akumulasi Penyusutan Peralatan",
    "2-10000": "Utang Usaha",
    "2-20000": "Utang Bank",
    "2-30000": "Pendapatan Diterima di Muka",
    "2-31000": "Utang Gaji",
    "3-30000": "Modal",
    "3-40000": "Prive",
    "4-40000": "Pendapatan",
    "5-50000": "HPP",
    "6-60000": "Beban Pakan",
    "6-60100": "Beban Listrik & Air",
    "6-60200": "Beban Gaji",
    "6-60300": "Beban Lain-lain",
    "6-60400": "Beban Penyusutan Kendaraan",
    "6-60500": "Beban Penyusutan Kandang",
    "6-60600": "Beban Penyusutan Peralatan",
    "6-60700": "Beban Perlengkapan"
}

# Custom CSS
st.markdown("""      
<style>
    .main, .block-container {
        background-color: #FDF3B9 ; 
    }
    [data-testid="stSidebar"] {
        background-color: #80644B; 
        padding-top: 0px;
    }
    [data-testid="stSidebar"] .stButton>button {
        background-color: #C8A574;
        color: white;
        border-radius: 8px;
        width: 100% ;
        text-align: left ;
        padding: 12px 16px ;
        margin: 4px 0 ;
    }

    /* Optional: tambahkan hover effect */
    [data-testid="stSidebar"] .stButton>button:hover {
        background-color: #B8945F;
        border-color: #A5834F;
    }
    .sidebar-title {
        color: #ffffff;      
        font-weight: 700;    
        font-size: 36px;     
    }
    .main-header {
        background: linear-gradient(135deg, #80644B 0%, #8B6F47 100%);
        padding: 1.5rem;
        border-radius: 10px;
        margin-bottom: 2rem;
        color: white;
    }
    .card {
        background: white;
        padding: 1.5rem;
        border-radius: 10px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        margin-bottom: 1rem;
    }
    .metric-card {
        background: linear-gradient(135deg, #DDB27A 0%, #C8A574 100%);
        padding: 1rem;
        border-radius: 8px;
        text-align: center;
        color: white;
    }
    .stButton>button {
        width: 100%;
        background-color: #8B6F47;
        color: white;
        border-radius: 8px;
        padding: 0.5rem 1rem;
        border: none;
    }
    .stButton>button:hover {
        background-color: #5A4634;
    }
    div[data-testid="stSidebar"] {
        background-color: #DDB27A;
    }     
</style>
""", unsafe_allow_html=True)

# Helper Functions
def format_rupiah(amount):
    """Format angka ke format rupiah"""
    try:
        return f"Rp {float(amount):,.0f}".replace(',', '.')
    except:
        return f"Rp 0"

def safe_parse_int_from_qtytext(qty_text):
    try:
        if qty_text is None:
            return 0
        if isinstance(qty_text, (int, float)):
            return int(qty_text)
        s = str(qty_text).strip()
        if s == "":
            return 0
        first = s.split()[0]
        return int(float(first))
    except:
        return 0

def safe_parse_price(text):
    try:
        if text is None:
            return 0.0
        if isinstance(text, (int, float)):
            return float(text)
        s = str(text)
        s = s.replace('Rp', '').replace(' ', '')
        s = s.replace('.', '').replace(',', '')
        return float(s) if s != "" else 0.0
    except:
        return 0.0
    
def get_inventory_account(product_name):
    """Mengembalikan akun persediaan berdasarkan nama produk"""
    product_to_account = {
        "Kerbau Dewasa Jantan": "1-12000 - Persediaan Kerbau Dewasa Jantan",
        "Kerbau Dewasa Betina": "1-12100 - Persediaan Kerbau Dewasa Betina", 
        "Kerbau Remaja Jantan": "1-12200 - Persediaan Kerbau Remaja Jantan",
        "Kerbau Remaja Betina": "1-12300 - Persediaan Kerbau Remaja Betina",
        "Anak Kerbau Jantan": "1-12400 - Persediaan Anak Kerbau Jantan",
        "Anak Kerbau Betina": "1-12500 - Persediaan Anak Kerbau Betina"
    }
    
    for key, account in product_to_account.items():
        if key.lower() in product_name.lower():
            return account
    
    return "1-12000 - Persediaan Kerbau Dewasa Jantan"

def delete_purchase_transaction(purchase_data):
    """Hapus transaksi pembelian dari semua sistem"""
    try:
        # 1. Hapus dari database pembelian
        wb = openpyxl.load_workbook('databasesia.xlsx')
        ws_purchases = wb['Purchases']
        
        # Hapus baris dari purchases
        ws_purchases.delete_rows(purchase_data['row_index'])
        
        # 2. Update inventory (kurangi stok dan hitung ulang average cost)
        ws_inventory = wb['Inventory']
        product_name = purchase_data['product_name']
        quantity_to_remove = safe_parse_int_from_qtytext(purchase_data['quantity'])
        price_to_remove = safe_parse_price(purchase_data['price'])
        total_to_remove = safe_parse_price(purchase_data['total'])
        
        # Cari produk di inventory
        for row in ws_inventory.iter_rows(min_row=2, values_only=False):
            if row[0].value and str(row[0].value).strip().lower() == product_name.strip().lower():
                current_qty_str = str(row[1].value) if row[1].value else "0"
                current_qty = safe_parse_int_from_qtytext(current_qty_str)
                current_avg_price = safe_parse_price(row[2].value) if row[2].value else 0
                current_total = safe_parse_price(row[3].value) if row[3].value else 0
                
                # Hitung quantity baru
                new_qty = current_qty - quantity_to_remove
                
                if new_qty <= 0:
                    # Hapus produk dari inventory jika stok habis
                    ws_inventory.delete_rows(row[0].row)
                else:
                    # Hitung average price baru
                    if new_qty > 0:
                        # Total value sebelum penghapusan
                        total_value_before = current_avg_price * current_qty
                        # Total value yang dihapus
                        total_value_removed = price_to_remove * quantity_to_remove
                        # Total value setelah penghapusan
                        total_value_after = total_value_before - total_value_removed
                        # Average price baru
                        new_avg_price = total_value_after / new_qty
                        
                        # Update inventory
                        unit = current_qty_str.split()[1] if len(current_qty_str.split()) > 1 else "unit"
                        row[1].value = f"{new_qty} {unit}"
                        row[2].value = round(new_avg_price, 2)
                        row[3].value = round(new_avg_price * new_qty, 2)
                
                break
        
        wb.save('databasesia.xlsx')
        wb.close()
        
        return True
        
    except Exception as e:
        st.error(f"Error dalam delete_purchase_transaction: {e}")
        return False

def delete_sales_transaction(sale_data):
    """Hapus transaksi penjualan dari semua sistem"""
    try:
        # 1. Hapus dari database penjualan
        wb = openpyxl.load_workbook('databasesia.xlsx')
        ws_sales = wb['Sales']
        
        # Hapus baris dari sales
        ws_sales.delete_rows(sale_data['row_index'])
        
        # 2. Update inventory (tambahkan kembali stok yang terjual)
        ws_inventory = wb['Inventory']
        product_name = sale_data['product_name']
        quantity_to_restore = safe_parse_int_from_qtytext(sale_data['quantity'])
        selling_price = safe_parse_price(sale_data['price'])
        
        # Cari produk di inventory untuk mendapatkan HPP
        hpp_price = 0
        product_found = False
        
        for row in ws_inventory.iter_rows(min_row=2, values_only=False):
            if row[0].value and str(row[0].value).strip().lower() == product_name.strip().lower():
                current_qty_str = str(row[1].value) if row[1].value else "0"
                current_qty = safe_parse_int_from_qtytext(current_qty_str)
                current_avg_price = safe_parse_price(row[2].value) if row[2].value else 0
                
                # Kembalikan stok
                new_qty = current_qty + quantity_to_restore
                unit = current_qty_str.split()[1] if len(current_qty_str.split()) > 1 else "ekor"
                
                # Hitung average price baru (gunakan harga average yang ada)
                new_avg_price = current_avg_price  # Tetap menggunakan average price yang ada
                
                # Update inventory
                row[1].value = f"{new_qty} {unit}"
                row[2].value = round(new_avg_price, 2)
                row[3].value = round(new_avg_price * new_qty, 2)
                
                hpp_price = current_avg_price
                product_found = True
                break
        
        # Jika produk tidak ditemukan, buat baru
        if not product_found:
            ws_inventory.append([
                product_name,
                f"{quantity_to_restore} ekor",
                hpp_price,
                hpp_price * quantity_to_restore
            ])
        
        wb.save('databasesia.xlsx')
        wb.close()
        
        return True
        
    except Exception as e:
        st.error(f"Error dalam delete_sales_transaction: {e}")
        return False

def create_workbook_if_not_exists():
    if not os.path.exists('databasesia.xlsx'):
        wb = openpyxl.Workbook()
        ws_inventory = wb.active
        ws_inventory.title = "Inventory"
        ws_inventory.append(['Product Name', 'Product Quantity', 'Product Price', 'Total Price'])
        ws_sales = wb.create_sheet("Sales")
        ws_sales.append(['Date', 'Product Name', 'Product Quantity', 'Product Price', 'Total Sales', 'Timestamp'])
        ws_purchases = wb.create_sheet("Purchases")
        ws_purchases.append(['Date', 'Product Name', 'Product Quantity', 'Product Price', 'Total Price', 'Timestamp'])
        wb.save('databasesia.xlsx')

create_workbook_if_not_exists()

# Initialize session state
if 'current_page' not in st.session_state:
    st.session_state.current_page = "Dashboard"
if 'order_list' not in st.session_state:
    st.session_state.order_list = []

# Sidebar Navigation
with st.sidebar:
    st.markdown('<p class="sidebar-title">🐃 BuffaBook</p>', unsafe_allow_html=True)
    st.markdown("---")
    
    menu_options = {
        "📊 Dashboard": "Dashboard",
        "📦 Kartu Persediaan": "Kartu Persediaan",
        "📈 Ringkasan Penjualan": "Ringkasan Penjualan",
        "✍️ Input Jurnal Umum": "Jurnal Umum",
        "📖 Jurnal Umum": "View Jurnal",
        "📚 Buku Besar": "Buku Besar",
        "⚖️ Neraca Saldo": "Neraca Saldo",
        "📘 Adjusting": "Jurnal Penyesuaian",
        "📋 Laporan Keuangan": "Laporan Keuangan"
    }
    
    for label, page in menu_options.items():
        if st.button(label, key=f"btn_{page}"):
            st.session_state.current_page = page
            st.rerun()


# Main Content Area
def show_dashboard():
    st.markdown('<div class="main-header"><h1>🐃 Welcome to BuffaBook</h1><p>Sistem Pencatatan Peternakan Kerbau</p></div>', unsafe_allow_html=True)
    
    st.markdown("""
    <div class="card">
    <h3>Tentang BuffaBook</h3>
    <p>Selamat datang di BuffaBook – Sistem Pencatatan Peternakan Kerbau.</p>
    <p>Di sini Anda dapat mengelola inventory, pembelian, dan penjualan produk.</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Quick Stats
    col1, col2, col3 = st.columns(3)
    
    try:
        wb = openpyxl.load_workbook('databasesia.xlsx')
        
        # Total Inventory Items
        ws_inv = wb['Inventory']
        total_items = ws_inv.max_row - 1
        
        # Total Purchases
        ws_purch = wb['Purchases']
        total_purchases = 0
        for row in ws_purch.iter_rows(min_row=2, values_only=True):
            if row and row[4]:
                total_purchases += safe_parse_price(row[4])
        
        # Total Sales
        ws_sales = wb['Sales']
        total_sales = 0
        for row in ws_sales.iter_rows(min_row=2, values_only=True):
            if row and row[4]:
                total_sales += safe_parse_price(row[4])
        
        wb.close()
        
        with col1:
            st.markdown(f"""
            <div class="metric-card">
                <h2>{total_items}</h2>
                <p>Item Persediaan</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
            <div class="metric-card">
                <h2>{format_rupiah(total_purchases)}</h2>
                <p>Total Pembelian</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            st.markdown(f"""
            <div class="metric-card">
                <h2>{format_rupiah(total_sales)}</h2>
                <p>Total Penjualan</p>
            </div>
            """, unsafe_allow_html=True)
    
    except Exception as e:
        st.error(f"Error loading stats: {e}")

def show_kartu_persediaan():
    st.markdown('<div class="main-header"><h1>📦 Kartu Persediaan</h1></div>', unsafe_allow_html=True)
    
    # Predefined selling prices
    SELLING_PRICE = {
        "Kerbau Dewasa Jantan": 30000000,
        "Kerbau Dewasa Betina": 27000000,
        "Kerbau Remaja Jantan": 20000000,
        "Kerbau Remaja Betina": 17000000,
        "Anak Kerbau Jantan": 15000000,
        "Anak Kerbau Betina": 12000000,
    }
    
    # Tab untuk navigasi
    tab1, tab2, tab3, tab4 = st.tabs(["📝 Pembelian", "💰 Penjualan", "📋 Riwayat Transaksi", "📦 Kartu Persediaan"])
    
    with tab1:
        st.markdown("### 🛒 Tambah Pembelian Baru")
        
        with st.form("form_pembelian"):
            col1, col2 = st.columns(2)
            
            with col1:
                date = st.date_input("Tanggal", datetime.now())
                product_name = st.text_input("Nama Produk")
                # TAMBAHAN: Pilihan metode pembayaran
                payment_method = st.selectbox("Metode Pembayaran", ["Tunai", "Kredit"])
                
            with col2:
                col_qty, col_unit = st.columns([2, 1])
                with col_qty:
                    quantity = st.number_input("Jumlah", min_value=1, value=1)
                with col_unit:
                    unit = st.selectbox("Unit", ["kg", "pcs", "karung", "ekor"])
                
                price = st.number_input("Harga per Unit (Rp)", min_value=0, value=0)
            
            submit = st.form_submit_button("✅ Submit Pembelian", use_container_width=True)
            
            if submit:
                if not product_name:
                    st.error("Nama produk tidak boleh kosong!")
                else:
                    try:
                        total_price = price * quantity
                        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        
                        wb = openpyxl.load_workbook('databasesia.xlsx')
                        ws_purchases = wb['Purchases']
                        ws_inventory = wb['Inventory']
                        
                        # Update inventory
                        product_found = False
                        for row in ws_inventory.iter_rows(min_row=2, values_only=False):
                            if row[0].value and str(row[0].value).strip().lower() == product_name.strip().lower():
                                existing_qty_str = str(row[1].value) if row[1].value is not None else "0"
                                try:
                                    parts = existing_qty_str.split()
                                    qty_lama = int(parts[0]) if parts else 0
                                    existing_unit = parts[1] if len(parts) > 1 else ''
                                except:
                                    qty_lama = 0
                                    existing_unit = ''
                                
                                harga_lama = float(row[2].value or 0)
                                qty_baru = quantity
                                harga_baru = float(price)
                                
                                total_qty = qty_lama + qty_baru
                                if total_qty == 0:
                                    harga_rata2 = harga_baru
                                else:
                                    harga_rata2 = ((qty_lama * harga_lama) + (qty_baru * harga_baru)) / total_qty
                                
                                row[1].value = f"{total_qty} {existing_unit or unit}".strip()
                                row[2].value = round(harga_rata2, 2)
                                row[3].value = round(harga_rata2 * total_qty, 2)
                                
                                product_found = True
                                break
                        
                        if not product_found:
                            ws_inventory.append([
                                product_name,
                                f"{quantity} {unit}",
                                round(price, 2),
                                round(total_price, 2)
                            ])
                        
                        # Add to purchases
                        ws_purchases.append([
                            date.strftime('%Y-%m-%d'),
                            product_name,
                            f"{quantity} {unit}",
                            round(price, 2),
                            round(total_price, 2),
                            timestamp,
                            payment_method  # TAMBAHAN: Simpan metode pembayaran
                        ])
                        
                        wb.save('databasesia.xlsx')
                        wb.close()
                        
                        # ========== OTOMATIS BUAT JURNAL UMUM ==========
                        create_journal_workbook()  # Pastikan file jurnal ada
                        wb_journal = load_workbook('journal_ledger.xlsx')
                        ws_journal = wb_journal['Jurnal Umum']
                        ws_ledger = wb_journal['Buku Besar']
                        
                        date_str = date.strftime('%Y-%m-%d')
                        keterangan = f"Pembelian {product_name} - {quantity} {unit}"
                        
                        # Tentukan akun persediaan berdasarkan nama produk
                        inventory_account = "1-12000 - Persediaan Kerbau Dewasa Jantan"  # default
                        
                        # Mapping nama produk ke akun persediaan
                        product_to_account = {
                            "Kerbau Dewasa Jantan": "1-12000 - Persediaan Kerbau Dewasa Jantan",
                            "Kerbau Dewasa Betina": "1-12100 - Persediaan Kerbau Dewasa Betina", 
                            "Kerbau Remaja Jantan": "1-12200 - Persediaan Kerbau Remaja Jantan",
                            "Kerbau Remaja Betina": "1-12300 - Persediaan Kerbau Remaja Betina",
                            "Anak Kerbau Jantan": "1-12400 - Persediaan Anak Kerbau Jantan",
                            "Anak Kerbau Betina": "1-12500 - Persediaan Anak Kerbau Betina",
                        }
                        
                        # Cari akun yang cocok
                        for key, account in product_to_account.items():
                            if key.lower() in product_name.lower():
                                inventory_account = account
                                break
                        
                        # Tentukan akun kredit berdasarkan metode pembayaran
                        if payment_method == "Tunai":
                            credit_account = "1-10000 - Kas"
                        else:  # Kredit
                            credit_account = "2-10000 - Utang Usaha"
                        
                        # Simpan ke Jurnal Umum
                        # Baris Debit: Persediaan
                        ws_journal.append([
                            date_str,
                            inventory_account,
                            total_price,  # Debit
                            0,  # Kredit
                            keterangan
                        ])
                        
                        # Baris Kredit: Kas/Utang
                        ws_journal.append([
                            "",  # Tanggal kosong
                            credit_account,
                            0,  # Debit  
                            total_price,  # Kredit
                            ""  # Keterangan kosong
                        ])
                        
                        # Baris kosong pemisah
                        ws_journal.append(["", "", "", "", ""])
                        
                        # Simpan ke Buku Besar
                        # Hitung saldo terakhir untuk setiap akun
                        current_balances = {}
                        for row in ws_ledger.iter_rows(min_row=2, values_only=True):
                            if row and row[0]:
                                account = row[0]
                                saldo = safe_parse_price(row[5]) if row[5] else 0
                                current_balances[account] = saldo
                        
                        # Update saldo akun persediaan (debit)
                        current_balance_inventory = current_balances.get(inventory_account, 0)
                        new_balance_inventory = current_balance_inventory + total_price
                        ws_ledger.append([
                            inventory_account,
                            date_str,
                            keterangan,
                            total_price,  # Debit
                            0,  # Kredit
                            new_balance_inventory
                        ])
                        
                        # Update saldo akun kas/utang (kredit)
                        current_balance_credit = current_balances.get(credit_account, 0)
                        if payment_method == "Tunai":
                            new_balance_credit = current_balance_credit - total_price
                        else:
                            new_balance_credit = current_balance_credit + total_price  # Utang bertambah
                        
                        ws_ledger.append([
                            credit_account,
                            date_str,
                            keterangan,
                            0,  # Debit
                            total_price,  # Kredit
                            new_balance_credit
                        ])
                        
                        wb_journal.save('journal_ledger.xlsx')
                        wb_journal.close()
                        # ========== END OTOMATIS JURNAL ==========
                        
                        st.success("✅ Produk berhasil ditambahkan ke persediaan dan jurnal dibuat otomatis!")
                        st.rerun()
                        
                    except Exception as e:
                        st.error(f"❌ Error: {e}")
        
        # Riwayat Pembelian
        st.markdown("### 📋 Riwayat Pembelian")
        try:
            wb = openpyxl.load_workbook('databasesia.xlsx')
            ws = wb['Purchases']
            
            data = []
            total_purchases = 0.0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                
                date, product_name, product_quantity, product_price, total_price = row[:5]
                product_price_val = safe_parse_price(product_price)
                total_price_val = safe_parse_price(total_price)
                total_purchases += total_price_val
                
                data.append({
                    'Tanggal': date,
                    'Nama Produk': product_name,
                    'Kuantitas': product_quantity,
                    'Harga Satuan': format_rupiah(product_price_val),
                    'Total Harga': format_rupiah(total_price_val)
                })
            
            wb.close()
            
            if data:
                df = pd.DataFrame(data)
                st.dataframe(df, use_container_width=True, hide_index=True)
                
                st.markdown(f"""
                <div style="background: #10b981; color: white; padding: 1rem; border-radius: 8px; text-align: center; margin-top: 1rem;">
                    <h3>💰 Total Pembelian: {format_rupiah(total_purchases)}</h3>
                </div>
                """, unsafe_allow_html=True)
            else:
                st.info("📝 Belum ada data pembelian")
        
        except Exception as e:
            st.error(f"❌ Error loading data: {e}")

        st.markdown("### 🗑️ Hapus Pembelian")
        
        try:
            wb = openpyxl.load_workbook('databasesia.xlsx')
            ws_purchases = wb['Purchases']
            
            # Ambil data pembelian untuk dropdown
            purchase_options = []
            purchase_details = {}
            
            for i, row in enumerate(ws_purchases.iter_rows(min_row=2, values_only=True), 2):
                if row and row[0]:
                    date, product_name, quantity, price, total, timestamp, payment_method = row[:7]
                    key = f"{date} - {product_name} - {quantity} - {format_rupiah(safe_parse_price(total))}"
                    purchase_options.append(key)
                    purchase_details[key] = {
                        'row_index': i,
                        'date': date,
                        'product_name': product_name,
                        'quantity': quantity,
                        'price': price,
                        'total': total,
                        'timestamp': timestamp,
                        'payment_method': payment_method
                    }
            
            wb.close()
            
            if purchase_options:
                selected_purchase = st.selectbox("Pilih Pembelian yang akan dihapus:", purchase_options)
                
                col1, col2 = st.columns([3, 1])
                with col2:
                    if st.button("🗑️ Hapus Pembelian Terpilih", type="secondary", use_container_width=True):
                        if selected_purchase:
                            purchase_data = purchase_details[selected_purchase]
                            if delete_purchase_transaction(purchase_data):
                                st.success("✅ Pembelian berhasil dihapus dari semua sistem!")
                                st.rerun()
                            else:
                                st.error("❌ Gagal menghapus pembelian!")
            else:
                st.info("📝 Tidak ada data pembelian untuk dihapus")
                
        except Exception as e:
            st.error(f"❌ Error: {e}")

    with tab2:
        st.markdown("### 💰 Tambah Penjualan Baru")
        
        with st.form("form_penjualan"):
            col1, col2 = st.columns(2)
            
            with col1:
                date = st.date_input("Tanggal Penjualan", datetime.now())
                product_name = st.selectbox("Nama Produk", list(SELLING_PRICE.keys()))
                # TAMBAHAN: Pilihan metode pembayaran
                payment_method = st.selectbox("Metode Pembayaran", ["Tunai", "Kredit"], key="sales_payment")
            
            with col2:
                quantity = st.number_input("Jumlah (ekor)", min_value=1, value=1, key="qty_sales")
                
                # Display price (readonly)
                selling_price = SELLING_PRICE[product_name]
                st.text_input("Harga Jual per Unit", value=format_rupiah(selling_price), disabled=True)
            
            add_to_list = st.form_submit_button("➕ Tambah ke Daftar", use_container_width=True)
            
            if add_to_list:
                try:
                    # Check inventory dan ambil HPP
                    wb = openpyxl.load_workbook('databasesia.xlsx')
                    ws_inv = wb['Inventory']
                    
                    stock_available = False
                    hpp_price = 0
                    for row in ws_inv.iter_rows(min_row=2, values_only=False):
                        if row[0].value and str(row[0].value).strip().lower() == product_name.strip().lower():
                            qty_str = str(row[1].value) if row[1].value is not None else "0"
                            try:
                                stock = int(qty_str.split()[0])
                            except:
                                stock = 0
                            
                            # Ambil HPP dari inventory
                            hpp_price = safe_parse_price(row[2].value) if row[2].value else 0
                            
                            if stock >= quantity:
                                # Update stock
                                unit = qty_str.split()[1] if len(qty_str.split()) > 1 else "ekor"
                                new_stock = stock - quantity
                                row[1].value = f"{new_stock} {unit}".strip()
                                
                                try:
                                    row[3].value = hpp_price * new_stock
                                except:
                                    pass
                                
                                stock_available = True
                            else:
                                st.error(f"❌ Stok {product_name} hanya {stock} ekor!")
                                wb.close()
                                break
                            break
                    
                    if stock_available:
                        wb.save('databasesia.xlsx')
                        wb.close()
                        
                        # Hitung total
                        total_sales = selling_price * quantity
                        total_hpp = hpp_price * quantity
                        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        
                        if 'order_list' not in st.session_state:
                            st.session_state.order_list = []
                        
                        # Simpan HPP juga di session state
                        st.session_state.order_list.append({
                            'date': date.strftime('%Y-%m-%d'),
                            'product_name': product_name,
                            'quantity': f"{quantity} ekor",
                            'price': selling_price,
                            'hpp_price': hpp_price,  # TAMBAHAN: Simpan HPP
                            'total': total_sales,
                            'total_hpp': total_hpp,  # TAMBAHAN: Simpan total HPP
                            'payment_method': payment_method,  # TAMBAHAN: Simpan metode pembayaran
                            'timestamp': timestamp
                        })
                        
                        st.success(f"✅ Penjualan {product_name} berhasil ditambahkan!")
                        st.rerun()
                    else:
                        if not stock_available:
                            st.error(f"❌ {product_name} tidak ditemukan di Inventory.")
                
                except Exception as e:
                    st.error(f"❌ Error: {e}")

        st.markdown("### 🛍️ Daftar Penjualan Sementara")
        
        if 'order_list' not in st.session_state:
            st.session_state.order_list = []
            
        if st.session_state.order_list:
            # Display order list
            order_data = []
            total_all_sales = 0
            total_all_hpp = 0
            for order in st.session_state.order_list:
                order_data.append({
                    'Tanggal': order['date'],
                    'Nama Produk': order['product_name'],
                    'Jumlah': order['quantity'],
                    'Harga Jual': format_rupiah(order['price']),
                    'HPP': format_rupiah(order['hpp_price']),
                    'Total': format_rupiah(order['total'])
                })
                total_all_sales += order['total']
                total_all_hpp += order['total_hpp']
            
            df = pd.DataFrame(order_data)
            st.dataframe(df, use_container_width=True, hide_index=True)
            
            col1, col2 = st.columns(2)
            with col1:
                st.markdown(f"""
                <div style="background: #3b82f6; color: white; padding: 1rem; border-radius: 8px; text-align: center; margin-top: 1rem;">
                    <h3>💰 Total Penjualan: {format_rupiah(total_all_sales)}</h3>
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                st.markdown(f"""
                <div style="background: #f59e0b; color: white; padding: 1rem; border-radius: 8px; text-align: center; margin-top: 1rem;">
                    <h3>📊 Total HPP: {format_rupiah(total_all_hpp)}</h3>
                </div>
                """, unsafe_allow_html=True)
            
            # Save all button - SEKARANG DI LUAR FORM
            col1, col2 = st.columns(2)
            with col1:
                if st.button("💾 Simpan Semua Penjualan", use_container_width=True):
                    try:
                        wb = openpyxl.load_workbook('databasesia.xlsx')
                        ws_sales = wb['Sales']
                        
                        for order in st.session_state.order_list:
                            ws_sales.append([
                                order['date'],
                                order['product_name'],
                                order['quantity'],
                                order['price'],
                                order['total'],
                                order['timestamp'],
                                order['payment_method']  # TAMBAHAN: Simpan metode pembayaran
                            ])
                        
                        wb.save('databasesia.xlsx')
                        wb.close()
                        
                        # ========== OTOMATIS BUAT JURNAL UNTUK SEMUA PENJUALAN ==========
                        create_journal_workbook()
                        wb_journal = load_workbook('journal_ledger.xlsx')
                        ws_journal = wb_journal['Jurnal Umum']
                        ws_ledger = wb_journal['Buku Besar']
                        
                        for order in st.session_state.order_list:
                            date_str = order['date']
                            product_name = order['product_name']
                            quantity = safe_parse_int_from_qtytext(order['quantity'])
                            total_sales = order['total']
                            total_hpp = order['total_hpp']
                            payment_method = order['payment_method']
                            keterangan = f"Penjualan {product_name} - {quantity} ekor"
                            
                            # Tentukan akun berdasarkan produk dan metode pembayaran
                            # Mapping nama produk ke akun persediaan
                            product_to_account = {
                                "Kerbau Dewasa Jantan": "1-12000 - Persediaan Kerbau Dewasa Jantan",
                                "Kerbau Dewasa Betina": "1-12100 - Persediaan Kerbau Dewasa Betina", 
                                "Kerbau Remaja Jantan": "1-12200 - Persediaan Kerbau Remaja Jantan",
                                "Kerbau Remaja Betina": "1-12300 - Persediaan Kerbau Remaja Betina",
                                "Anak Kerbau Jantan": "1-12400 - Persediaan Anak Kerbau Jantan",
                                "Anak Kerbau Betina": "1-12500 - Persediaan Anak Kerbau Betina"
                            }
                            
                            inventory_account = product_to_account.get(product_name, "1-12000 - Persediaan Kerbau Dewasa Jantan")
                            
                            # Tentukan akun debit berdasarkan metode pembayaran
                            if payment_method == "Tunai":
                                debit_account_1 = "1-10000 - Kas"
                            else:  # Kredit
                                debit_account_1 = "1-11000 - Piutang"
                            
                            # Hitung saldo terakhir
                            current_balances = {}
                            for row in ws_ledger.iter_rows(min_row=2, values_only=True):
                                if row and row[0]:
                                    account = row[0]
                                    saldo = safe_parse_price(row[5]) if row[5] else 0
                                    current_balances[account] = saldo
                            
                            ws_journal.append([
                                date_str,
                                debit_account_1,
                                total_sales,  # Debit
                                0,  # Kredit
                                keterangan
                            ])
                            
                            # Baris 2: Debit HPP
                            ws_journal.append([
                                "",
                                "5-50000 - HPP",
                                total_hpp,  # Debit
                                0,  # Kredit
                                ""
                            ])
                            
                            # Baris 3: Kredit Pendapatan
                            ws_journal.append([
                                "",
                                "4-40000 - Pendapatan",
                                0,  # Debit
                                total_sales,  # Kredit
                                ""
                            ])
                            
                            # Baris 4: Kredit Persediaan
                            ws_journal.append([
                                "",
                                inventory_account,
                                0,  # Debit
                                total_hpp,  # Kredit  
                                ""
                            ])
                            
                            # Baris kosong pemisah
                            ws_journal.append(["", "", "", "", ""])
                            
                            # UPDATE BUKU BESAR
                            # 1. Kas/Piutang (debit)
                            current_balance_debit1 = current_balances.get(debit_account_1, 0)
                            new_balance_debit1 = current_balance_debit1 + total_sales
                            ws_ledger.append([
                                debit_account_1,
                                date_str,
                                keterangan,
                                total_sales,  # Debit
                                0,  # Kredit
                                new_balance_debit1
                            ])
                            
                            # 2. HPP (debit)
                            current_balance_hpp = current_balances.get("5-50000 - HPP", 0)
                            new_balance_hpp = current_balance_hpp + total_hpp
                            ws_ledger.append([
                                "5-50000 - HPP",
                                date_str,
                                keterangan,
                                total_hpp,  # Debit
                                0,  # Kredit
                                new_balance_hpp
                            ])
                            
                            # 3. Pendapatan (kredit)
                            current_balance_pendapatan = current_balances.get("4-40000 - Pendapatan", 0)
                            new_balance_pendapatan = current_balance_pendapatan - total_sales  # Pendapatan normal balance kredit
                            ws_ledger.append([
                                "4-40000 - Pendapatan",
                                date_str,
                                keterangan,
                                0,  # Debit
                                total_sales,  # Kredit
                                new_balance_pendapatan
                            ])
                            
                            # 4. Persediaan (kredit)
                            current_balance_inventory = current_balances.get(inventory_account, 0)
                            new_balance_inventory = current_balance_inventory - total_hpp
                            ws_ledger.append([
                                inventory_account,
                                date_str,
                                keterangan,
                                0,  # Debit
                                total_hpp,  # Kredit
                                new_balance_inventory
                            ])
                        
                        wb_journal.save('journal_ledger.xlsx')
                        wb_journal.close()
                        # ========== END OTOMATIS JURNAL ==========
                        
                        st.session_state.order_list = []
                        st.success("✅ Semua penjualan berhasil disimpan dan jurnal dibuat otomatis!")
                        st.rerun()
                    
                    except Exception as e:
                        st.error(f"❌ Error: {e}")
            
            with col2:
                if st.button("🗑️ Hapus Semua Penjualan", use_container_width=True):
                    st.session_state.order_list = []
                    st.success("✅ Daftar penjualan berhasil dihapus!")
                    st.rerun()
        else:
            st.info("📝 Belum ada penjualan dalam daftar")

        st.markdown("### 🗑️ Hapus Penjualan")
        
        try:
            wb = openpyxl.load_workbook('databasesia.xlsx')
            ws_sales = wb['Sales']
            
            # Ambil data penjualan untuk dropdown
            sales_options = []
            sales_details = {}
            
            for i, row in enumerate(ws_sales.iter_rows(min_row=2, values_only=True), 2):
                if row and row[0]:
                    date, product_name, quantity, price, total, timestamp, payment_method = row[:7]
                    key = f"{date} - {product_name} - {quantity} - {format_rupiah(safe_parse_price(total))}"
                    sales_options.append(key)
                    sales_details[key] = {
                        'row_index': i,
                        'date': date,
                        'product_name': product_name,
                        'quantity': quantity,
                        'price': price,
                        'total': total,
                        'timestamp': timestamp,
                        'payment_method': payment_method
                    }
            
            wb.close()
            
            if sales_options:
                selected_sale = st.selectbox("Pilih Penjualan yang akan dihapus:", sales_options)
                
                col1, col2 = st.columns([3, 1])
                with col2:
                    if st.button("🗑️ Hapus Penjualan Terpilih", type="secondary", use_container_width=True):
                        if selected_sale:
                            sale_data = sales_details[selected_sale]
                            if delete_sales_transaction(sale_data):
                                st.success("✅ Penjualan berhasil dihapus dari semua sistem!")
                                st.rerun()
                            else:
                                st.error("❌ Gagal menghapus penjualan!")
            else:
                st.info("📝 Tidak ada data penjualan untuk dihapus")

        except Exception as e:
            st.error(f"❌ Error: {e}")

    with tab3:
        st.markdown("### 📋 Riwayat Transaksi Lengkap")
        
        try:
            wb = openpyxl.load_workbook('databasesia.xlsx')
            ws_purchases = wb['Purchases']
            ws_sales = wb['Sales']
            
            # Gabungkan data pembelian dan penjualan
            all_transactions = []
            
            # Data pembelian
            for row in ws_purchases.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                date, product_name, quantity, price, total, timestamp = row[:6]
                all_transactions.append({
                    'Tanggal': date,
                    'Tipe': 'Pembelian',
                    'Produk': product_name,
                    'Kuantitas': quantity,
                    'Harga/Unit': format_rupiah(safe_parse_price(price)),
                    'Total': format_rupiah(safe_parse_price(total)),
                    'Waktu': timestamp.split(' ')[1] if timestamp and ' ' in timestamp else ''
                })
            
            # Data penjualan
            for row in ws_sales.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                date, product_name, quantity, price, total, timestamp = row[:6]
                all_transactions.append({
                    'Tanggal': date,
                    'Tipe': 'Penjualan',
                    'Produk': product_name,
                    'Kuantitas': quantity,
                    'Harga/Unit': format_rupiah(safe_parse_price(price)),
                    'Total': format_rupiah(safe_parse_price(total)),
                    'Waktu': timestamp.split(' ')[1] if timestamp and ' ' in timestamp else ''
                })
            
            wb.close()
            
            if all_transactions:
                # Urutkan berdasarkan tanggal dan waktu
                all_transactions.sort(key=lambda x: (x['Tanggal'], x['Waktu']), reverse=True)
                
                df = pd.DataFrame(all_transactions)
                st.dataframe(df, use_container_width=True, hide_index=True)
                
                # Summary
                total_pembelian = sum(safe_parse_price(t['Total'].replace('Rp', '').replace('.', '').replace(',', '.').strip()) 
                                    for t in all_transactions if t['Tipe'] == 'Pembelian')
                total_penjualan = sum(safe_parse_price(t['Total'].replace('Rp', '').replace('.', '').replace(',', '.').strip()) 
                                    for t in all_transactions if t['Tipe'] == 'Penjualan')
                
                col1, col2 = st.columns(2)
                with col1:
                    st.metric("Total Pembelian", format_rupiah(total_pembelian))
                with col2:
                    st.metric("Total Penjualan", format_rupiah(total_penjualan))
                
            else:
                st.info("📝 Belum ada data transaksi")
        
        except Exception as e:
            st.error(f"❌ Error loading transaction data: {e}")

    with tab4:
        st.markdown("### 📊 Kartu Persediaan Detail")
        
        try:
            wb = openpyxl.load_workbook('databasesia.xlsx')
            ws_inventory = wb['Inventory']
            ws_purchases = wb['Purchases']
            ws_sales = wb['Sales']
            
            # Ambil semua produk dari inventory
            products = []
            for row in ws_inventory.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    product_name = row[0]
                    quantity_balance = row[1]  # Format: "10 kg"
                    price_balance = safe_parse_price(row[2])
                    total_balance = safe_parse_price(row[3])
                    
                    # Parse quantity balance
                    qty_balance = safe_parse_int_from_qtytext(quantity_balance)
                    unit_balance = str(quantity_balance).replace(str(qty_balance), "").strip()
                    
                    # Ambil data pembelian untuk produk ini
                    purchases_data = []
                    for purchase_row in ws_purchases.iter_rows(min_row=2, values_only=True):
                        if purchase_row and purchase_row[1] and str(purchase_row[1]).strip().lower() == product_name.lower():
                            date = purchase_row[0]
                            qty_purchase = safe_parse_int_from_qtytext(purchase_row[2])
                            price_purchase = safe_parse_price(purchase_row[3])
                            total_purchase = safe_parse_price(purchase_row[4])
                            timestamp = purchase_row[5] if len(purchase_row) > 5 else ""
                            
                            purchases_data.append({
                                'tanggal': date,
                                'timestamp': timestamp,
                                'qty': qty_purchase,
                                'price': price_purchase,
                                'total': total_purchase
                            })
                    
                    # Ambil data penjualan untuk produk ini
                    sales_data = []
                    for sales_row in ws_sales.iter_rows(min_row=2, values_only=True):
                        if sales_row and sales_row[1] and str(sales_row[1]).strip().lower() == product_name.lower():
                            date = sales_row[0]
                            qty_sales = safe_parse_int_from_qtytext(sales_row[2])
                            price_sales = safe_parse_price(sales_row[3])
                            total_sales = safe_parse_price(sales_row[4])
                            timestamp = sales_row[5] if len(sales_row) > 5 else ""
                            
                            sales_data.append({
                                'tanggal': date,
                                'timestamp': timestamp,
                                'qty': qty_sales,
                                'price': price_sales,
                                'total': total_sales
                            })
                    
                    # Gabungkan semua transaksi
                    all_transactions = []
                    
                    # Tambahkan pembelian
                    for purchase in purchases_data:
                        all_transactions.append({
                            'tanggal': purchase['tanggal'],
                            'timestamp': purchase['timestamp'],
                            'type': 'Pembelian',
                            'qty': purchase['qty'],
                            'price': purchase['price'],
                            'total': purchase['total']
                        })
                    
                    # Tambahkan penjualan
                    for sale in sales_data:
                        all_transactions.append({
                            'tanggal': sale['tanggal'],
                            'timestamp': sale['timestamp'],
                            'type': 'Penjualan',
                            'qty': -sale['qty'],  # Negative untuk penjualan
                            'price': sale['price'],
                            'total': sale['total']
                        })
                    
                    # Urutkan berdasarkan tanggal DAN timestamp
                    all_transactions.sort(key=lambda x: (
                        x['tanggal'] if x['tanggal'] else "0000-00-00",
                        x['timestamp'] if x['timestamp'] else "00:00:00"
                    ))
                    
                    # Hitung running balance dengan metode average
                    running_qty = 0
                    running_avg_price = 0
                    running_total = 0
                    
                    transaction_details = []
                    
                    for trans in all_transactions:
                        if trans['type'] == 'Pembelian':
                            # Metode Average
                            if running_qty == 0:
                                running_avg_price = trans['price']
                            else:
                                total_value_before = running_qty * running_avg_price
                                total_value_new = trans['qty'] * trans['price']
                                running_avg_price = (total_value_before + total_value_new) / (running_qty + trans['qty'])
                            
                            running_qty += trans['qty']
                            running_total = running_qty * running_avg_price
                        
                        elif trans['type'] == 'Penjualan':
                            # Untuk penjualan, harga pakai average price yang ada
                            running_qty += trans['qty']  # trans['qty'] sudah negative
                            running_total = running_qty * running_avg_price
                        
                        # Format timestamp untuk display
                        display_time = ""
                        if trans['timestamp']:
                            try:
                                if ' ' in trans['timestamp']:
                                    date_part, time_part = trans['timestamp'].split(' ')
                                    display_time = time_part
                                else:
                                    display_time = trans['timestamp']
                            except:
                                display_time = trans['timestamp']
                        
                        transaction_details.append({
                            'Tanggal': trans['tanggal'],
                            'Waktu': display_time,
                            'Tipe': trans['type'],
                            'Qty_Pembelian': trans['qty'] if trans['type'] == 'Pembelian' else 0,
                            'Harga_Pembelian': format_rupiah(trans['price']) if trans['type'] == 'Pembelian' else '',
                            'Total_Pembelian': format_rupiah(trans['total']) if trans['type'] == 'Pembelian' else '',
                            'Qty_Penjualan': abs(trans['qty']) if trans['type'] == 'Penjualan' else 0,
                            'Harga_Penjualan': format_rupiah(trans['price']) if trans['type'] == 'Penjualan' else '',
                            'Total_Penjualan': format_rupiah(trans['total']) if trans['type'] == 'Penjualan' else '',
                            'Qty_Balance': running_qty,
                            'Harga_Balance': format_rupiah(running_avg_price),
                            'Total_Balance': format_rupiah(running_total)
                        })
                    
                    # Tambahkan balance akhir jika ada transaksi
                    if not transaction_details and qty_balance > 0:
                        transaction_details.append({
                            'Tanggal': '-',
                            'Waktu': '',
                            'Tipe': 'Balance Awal',
                            'Qty_Pembelian': 0,
                            'Harga_Pembelian': '',
                            'Total_Pembelian': '',
                            'Qty_Penjualan': 0,
                            'Harga_Penjualan': '',
                            'Total_Penjualan': '',
                            'Qty_Balance': qty_balance,
                            'Harga_Balance': format_rupiah(price_balance),
                            'Total_Balance': format_rupiah(total_balance)
                        })
                    
                    products.append({
                        'product_name': product_name,
                        'unit': unit_balance,
                        'current_stock': qty_balance,
                        'current_value': total_balance,
                        'transactions': transaction_details
                    })
            
            wb.close()
            
            # Tampilkan summary persediaan
            st.markdown("### 📈 Summary Persediaan")
            if products:
                summary_data = []
                total_inventory_value = 0
                
                for product in products:
                    summary_data.append({
                        'Nama Produk': product['product_name'],
                        'Stok Saat Ini': f"{product['current_stock']} {product['unit']}",
                        'Nilai Persediaan': format_rupiah(product['current_value'])
                    })
                    total_inventory_value += product['current_value']
                
                df_summary = pd.DataFrame(summary_data)
                st.dataframe(df_summary, use_container_width=True, hide_index=True)
                
                st.metric("Total Nilai Persediaan", format_rupiah(total_inventory_value))
                
                # Tampilkan detail kartu persediaan per produk
                st.markdown("### 📋 Detail Kartu Persediaan per Produk")
                for product in products:
                    with st.expander(f"📦 {product['product_name']} ({product['unit']}) - Stok: {product['current_stock']}"):
                        if product['transactions']:
                            df_detail = pd.DataFrame(product['transactions'])
                            
                            # Rename kolom untuk tampilan yang lebih baik
                            df_display = df_detail.rename(columns={
                                'Qty_Pembelian': 'Qty Pembelian',
                                'Harga_Pembelian': 'Harga/Unit Pembelian', 
                                'Total_Pembelian': 'Total Pembelian',
                                'Qty_Penjualan': 'Qty Penjualan',
                                'Harga_Penjualan': 'Harga/Unit Penjualan',
                                'Total_Penjualan': 'Total Penjualan',
                                'Qty_Balance': 'Qty Balance',
                                'Harga_Balance': 'Harga/Unit Balance',
                                'Total_Balance': 'Total Balance'
                            })
                            
                            st.dataframe(df_display, use_container_width=True, hide_index=True)
                        else:
                            st.info("Belum ada transaksi untuk produk ini")
            else:
                st.info("📝 Belum ada data persediaan")
        
        except Exception as e:
            st.error(f"❌ Error loading inventory data: {e}")

# Fungsi helper (pastikan fungsi-fungsi ini ada)
def safe_parse_price(price_value):
    """Parse price value safely"""
    if price_value is None:
        return 0.0
    try:
        if isinstance(price_value, (int, float)):
            return float(price_value)
        price_str = str(price_value).replace('Rp', '').replace('.', '').replace(',', '.').strip()
        return float(price_str) if price_str else 0.0
    except:
        return 0.0

def safe_parse_int_from_qtytext(qty_text):
    """Parse integer from quantity text like '10 kg'"""
    if qty_text is None:
        return 0
    try:
        if isinstance(qty_text, (int, float)):
            return int(qty_text)
        parts = str(qty_text).split()
        return int(parts[0]) if parts else 0
    except:
        return 0

def format_rupiah(amount):
    """Format number to Rupiah currency"""
    try:
        if amount == 0:
            return "Rp 0"
        return f"Rp {amount:,.0f}".replace(",", ".")
    except:
        return "Rp 0"

def show_ringkasan_penjualan():
    st.markdown('<div class="main-header"><h1>📈 Ringkasan Penjualan</h1></div>', unsafe_allow_html=True)
    
    try:
        wb = openpyxl.load_workbook('databasesia.xlsx')
        ws_sales = wb['Sales']
        ws_inventory = wb['Inventory']
        
        # Get HPP from inventory
        HPP_dict = {}
        for row in ws_inventory.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                product_name = row[0]
                HPP_price = safe_parse_price(row[2])
                HPP_dict[product_name] = HPP_price
        
        # Process sales data
        data = []
        total_income_all = 0.0
        total_HPP_all = 0.0
        total_profit_all = 0.0
        
        for row in ws_sales.iter_rows(min_row=2, values_only=True):
            if not row or row[1] is None:
                continue
            
            date = row[0]
            product_name = row[1]
            qty_text = row[2]
            selling_price_unit = safe_parse_price(row[3])
            total_income = safe_parse_price(row[4])
            
            qty = safe_parse_int_from_qtytext(qty_text)
            HPP_unit = HPP_dict.get(product_name, 0)
            total_HPP = HPP_unit * qty
            gross_profit = total_income - total_HPP
            
            total_income_all += total_income
            total_HPP_all += total_HPP
            total_profit_all += gross_profit
            
            data.append({
                'Tanggal': date,
                'Produk': product_name,
                'Qty': qty,
                'Harga Jual': format_rupiah(selling_price_unit),
                'HPP/Unit': format_rupiah(HPP_unit),
                'Total Income': format_rupiah(total_income),
                'Total HPP': format_rupiah(total_HPP),
                'Gross Profit': format_rupiah(gross_profit)
            })
        
        wb.close()
        
        if data:
            df = pd.DataFrame(data)
            st.dataframe(df, use_container_width=True, hide_index=True)
            
            # Summary
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.markdown(f"""
                <div style="background: #10b981; color: white; padding: 1rem; border-radius: 8px; text-align: center;">
                    <h4>Total Income</h4>
                    <h2>{format_rupiah(total_income_all)}</h2>
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                st.markdown(f"""
                <div style="background: #f59e0b; color: white; padding: 1rem; border-radius: 8px; text-align: center;">
                    <h4>Total HPP</h4>
                    <h2>{format_rupiah(total_HPP_all)}</h2>
                </div>
                """, unsafe_allow_html=True)
            
            with col3:
                profit_color = "#10b981" if total_profit_all > 0 else "#ef4444"
                st.markdown(f"""
                <div style="background: {profit_color}; color: white; padding: 1rem; border-radius: 8px; text-align: center;">
                    <h4>Profit</h4>
                    <h2>{format_rupiah(total_profit_all)}</h2>
                </div>
                """, unsafe_allow_html=True)
        else:
            st.info("Belum ada data penjualan")
    
    except Exception as e:
        st.error(f"Error: {e}")

def create_journal_workbook():
    """Create journal_ledger.xlsx if not exists"""
    if not os.path.exists('journal_ledger.xlsx'):
        wb = Workbook()
        
        # Jurnal Umum sheet
        ws_journal = wb.active
        ws_journal.title = "Jurnal Umum"
        ws_journal.append(["Tanggal", "Akun", "Debit", "Kredit", "Keterangan"])
        
        # Buku Besar sheet
        ws_ledger = wb.create_sheet("Buku Besar")
        ws_ledger.append(["Akun", "Tanggal", "Keterangan", "Debit", "Kredit", "Saldo"])
        
        wb.save('journal_ledger.xlsx')

def show_jurnal_umum():
    st.markdown('<div class="main-header"><h1>📒 Input Jurnal Umum</h1></div>', unsafe_allow_html=True)
    
    create_journal_workbook()
    
    st.markdown("### Input Transaksi Baru")
    
    # Initialize session state untuk dynamic form
    if 'debit_accounts' not in st.session_state:
        st.session_state.debit_accounts = [{'account': '', 'amount': 0}]
    if 'credit_accounts' not in st.session_state:
        st.session_state.credit_accounts = [{'account': '', 'amount': 0}]
    
    # Tombol tambah akun di LUAR form
    col1, col2 = st.columns(2)
    with col1:
        if st.button("➕ Tambah Akun Debit", key="add_debit"):
            st.session_state.debit_accounts.append({'account': '', 'amount': 0})
            st.rerun()
    with col2:
        if st.button("➕ Tambah Akun Kredit", key="add_credit"):
            st.session_state.credit_accounts.append({'account': '', 'amount': 0})
            st.rerun()
    
    with st.form("form_jurnal", clear_on_submit=True):
        date = st.date_input("Tanggal", datetime.now())
        
        st.markdown("**Akun Debit**")
        
        # Display debit accounts
        for i in range(len(st.session_state.debit_accounts)):
            st.markdown(f"**Debit {i+1}**")
            col_acc, col_amt, col_del = st.columns([2, 1, 0.5])
            with col_acc:
                account_options = [f"{code} - {name}" for code, name in ACCOUNTS.items()]
                # Cari index akun yang dipilih
                current_account = st.session_state.debit_accounts[i]['account']
                default_index = account_options.index(current_account) if current_account in account_options else 0
                
                selected_account = st.selectbox(
                    f"Pilih Akun Debit {i+1}", 
                    account_options, 
                    key=f"debit_acc_{i}",
                    index=default_index
                )
                st.session_state.debit_accounts[i]['account'] = selected_account
            with col_amt:
                amount = st.number_input(
                    f"Nominal Debit {i+1}", 
                    min_value=0, 
                    value=st.session_state.debit_accounts[i]['amount'],
                    key=f"debit_amt_{i}"
                )
                st.session_state.debit_accounts[i]['amount'] = amount
            with col_del:
                if i > 0:  # Hanya tampilkan tombol hapus untuk akun tambahan
                    # Gunakan form_submit_button dengan type secondary untuk tombol hapus
                    if st.form_submit_button("❌", type="secondary", key=f"del_debit_{i}"):
                        st.session_state.debit_accounts.pop(i)
                        st.rerun()
        
        st.markdown("---")
        st.markdown("**Akun Kredit**")
        
        # Display credit accounts
        for i in range(len(st.session_state.credit_accounts)):
            st.markdown(f"**Kredit {i+1}**")
            col_acc, col_amt, col_del = st.columns([2, 1, 0.5])
            with col_acc:
                account_options = [f"{code} - {name}" for code, name in ACCOUNTS.items()]
                # Cari index akun yang dipilih
                current_account = st.session_state.credit_accounts[i]['account']
                default_index = account_options.index(current_account) if current_account in account_options else 0
                
                selected_account = st.selectbox(
                    f"Pilih Akun Kredit {i+1}", 
                    account_options, 
                    key=f"credit_acc_{i}",
                    index=default_index
                )
                st.session_state.credit_accounts[i]['account'] = selected_account
            with col_amt:
                amount = st.number_input(
                    f"Nominal Kredit {i+1}", 
                    min_value=0, 
                    value=st.session_state.credit_accounts[i]['amount'],
                    key=f"credit_amt_{i}"
                )
                st.session_state.credit_accounts[i]['amount'] = amount
            with col_del:
                if i > 0:  # Hanya tampilkan tombol hapus untuk akun tambahan
                    # Gunakan form_submit_button dengan type secondary untuk tombol hapus
                    if st.form_submit_button("❌", type="secondary", key=f"del_credit_{i}"):
                        st.session_state.credit_accounts.pop(i)
                        st.rerun()
        
        st.markdown("---")
        keterangan = st.text_input("Keterangan Transaksi", placeholder="Contoh: Pembelian perlengkapan kantor")
        
        submit = st.form_submit_button("💾 Simpan Jurnal", use_container_width=True)
        
        if submit:
            # Filter hanya akun yang memiliki nominal > 0
            valid_debit_accounts = [acc for acc in st.session_state.debit_accounts if acc['amount'] > 0 and acc['account']]
            valid_credit_accounts = [acc for acc in st.session_state.credit_accounts if acc['amount'] > 0 and acc['account']]
            
            # Validation
            total_debit = sum(item['amount'] for item in valid_debit_accounts)
            total_credit = sum(item['amount'] for item in valid_credit_accounts)
            
            if total_debit == 0 or total_credit == 0:
                st.error("Nominal debit dan kredit harus lebih dari 0")
            elif total_debit != total_credit:
                st.error(f"**Tidak Balance!** Total Debit ({format_rupiah(total_debit)}) ≠ Total Kredit ({format_rupiah(total_credit)})")
            elif not valid_debit_accounts or not valid_credit_accounts:
                st.error("Harap isi minimal satu akun debit dan satu akun kredit")
            else:
                try:
                    wb = load_workbook('journal_ledger.xlsx')
                    ws_journal = wb['Jurnal Umum']
                    ws_ledger = wb['Buku Besar']
                    
                    date_str = date.strftime('%Y-%m-%d')
                    
                    # 1. SIMPAN KE JURNAL UMUM
                    # Baris pertama: akun debit pertama dengan keterangan
                    if valid_debit_accounts:
                        first_debit = valid_debit_accounts[0]
                        ws_journal.append([
                            date_str,
                            first_debit['account'],
                            first_debit['amount'],  # Debit
                            0,  # Kredit = 0 untuk akun debit
                            keterangan  # Keterangan hanya di baris pertama
                        ])
                    
                    # Baris untuk akun debit lainnya (tanpa keterangan)
                    for i in range(1, len(valid_debit_accounts)):
                        debit = valid_debit_accounts[i]
                        ws_journal.append([
                            "",  # Tanggal kosong
                            debit['account'],
                            debit['amount'],  # Debit
                            0,  # Kredit = 0
                            ""  # Keterangan kosong
                        ])
                    
                    # Baris untuk akun kredit (semua tanpa keterangan)
                    for credit in valid_credit_accounts:
                        ws_journal.append([
                            "",  # Tanggal kosong
                            credit['account'],
                            0,  # Debit = 0 untuk akun kredit
                            credit['amount'],  # Kredit
                            ""  # Keterangan kosong
                        ])
                    
                    # Tambah baris kosong untuk pemisah antar transaksi
                    ws_journal.append(["", "", "", "", ""])
                    
                    # 2. SIMPAN KE BUKU BESAR (LEDGER)
                    # Hitung saldo akhir untuk setiap akun sebelum transaksi ini
                    current_balances = {}
                    for row in ws_ledger.iter_rows(min_row=2, values_only=True):
                        if row and row[0]:  # Jika ada akun
                            account = row[0]
                            debit = safe_parse_price(row[3]) if row[3] else 0
                            kredit = safe_parse_price(row[4]) if row[4] else 0
                            saldo = safe_parse_price(row[5]) if row[5] else 0
                            
                            if account not in current_balances:
                                current_balances[account] = saldo
                    
                    # Catat transaksi debit ke Buku Besar
                    for debit in valid_debit_accounts:
                        account = debit['account']
                        nominal = debit['amount']
                        
                        # Hitung saldo baru
                        current_balance = current_balances.get(account, 0)
                        new_balance = current_balance + nominal
                        current_balances[account] = new_balance
                        
                        ws_ledger.append([
                            account,
                            date_str,
                            keterangan,
                            nominal,  # Debit
                            0,  # Kredit = 0
                            new_balance
                        ])
                    
                    # Catat transaksi kredit ke Buku Besar
                    for credit in valid_credit_accounts:
                        account = credit['account']
                        nominal = credit['amount']
                        
                        # Hitung saldo baru
                        current_balance = current_balances.get(account, 0)
                        new_balance = current_balance - nominal
                        current_balances[account] = new_balance
                        
                        ws_ledger.append([
                            account,
                            date_str,
                            keterangan,
                            0,  # Debit = 0
                            nominal,  # Kredit
                            new_balance
                        ])
                    
                    wb.save('journal_ledger.xlsx')
                    wb.close()
                    
                    st.success("✅ Jurnal berhasil disimpan ke Jurnal Umum dan Buku Besar!")
                    
                    # Reset form setelah berhasil simpan
                    st.session_state.debit_accounts = [{'account': '', 'amount': 0}]
                    st.session_state.credit_accounts = [{'account': '', 'amount': 0}]
                    st.rerun()
                
                except Exception as e:
                    st.error(f"Error: {e}")

def show_view_jurnal():
    st.markdown('<div class="main-header"><h1>📖 Lihat Jurnal Umum</h1></div>', unsafe_allow_html=True)

    try:
        wb = load_workbook('journal_ledger.xlsx')
        ws = wb['Jurnal Umum']

        data = []
        row_indices = []  # Simpan index baris untuk referensi hapus
        transaction_groups = {}   # Kelompokkan transaksi
        
        # Baca semua data dulu
        all_rows = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if any(row[:4]):  # Skip baris yang benar-benar kosong
                tanggal, akun, debit, kredit, keterangan = row
                all_rows.append({
                    'row_index': i,
                    'tanggal': tanggal or '',
                    'akun': akun or '',
                    'debit': debit,
                    'kredit': kredit,
                    'keterangan': keterangan or ''
                })

        # Kelompokkan transaksi berdasarkan baris kosong
        current_group = []
        groups = []
        
        for row in all_rows:
            # Jika baris memiliki tanggal (bukan kosong), ini kemungkinan awal transaksi
            if row['tanggal']:
                # Simpan group sebelumnya jika ada
                if current_group:
                    groups.append(current_group.copy())
                    current_group = []
            
            current_group.append(row)
        
        # Simpan group terakhir
        if current_group:
            groups.append(current_group)

        # Format data untuk display dan hapus
        group_counter = 0
        for group in groups:
            if group:  # Pastikan group tidak kosong
                group_tanggal = group[0]['tanggal'] if group[0]['tanggal'] else ''
                group_keterangan = group[0]['keterangan'] if group[0]['keterangan'] else 'Transaksi tanpa keterangan'
                
                transaction_groups[group_counter] = {
                    'keterangan': group_keterangan,
                    'tanggal': group_tanggal,
                    'rows': group
                }
                
                # Tambahkan ke data untuk display
                for entry in group:
                    data.append({
                        'No': len(data) + 1,
                        'Tanggal': entry['tanggal'],
                        'Akun': entry['akun'],
                        'Debit': format_rupiah(entry['debit']) if entry['debit'] else '',
                        'Kredit': format_rupiah(entry['kredit']) if entry['kredit'] else '',
                        'Keterangan': entry['keterangan'],
                        'RowIndex': entry['row_index'],
                        'GroupID': group_counter
                    })
                    row_indices.append(entry['row_index'])
                
                group_counter += 1

        # Hitung total before closing workbook
        total_debit = sum(safe_parse_price(row[2]) for row in ws.iter_rows(min_row=2, values_only=True) if row and row[2])
        total_kredit = sum(safe_parse_price(row[3]) for row in ws.iter_rows(min_row=2, values_only=True) if row and row[3])

        wb.close()

        if data:
            df = pd.DataFrame(data)
            
            # ========== PERUBAHAN DI SINI ==========
            # Buat dataframe untuk display tanpa kolom No, RowIndex, GroupID
            df_display = df[['Tanggal', 'Akun', 'Debit', 'Kredit', 'Keterangan']]
            
            # Tampilkan dataframe hanya dengan kolom yang diinginkan
            st.dataframe(
                df_display,
                use_container_width=True,
                hide_index=True,
                height=600
            )

            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total Debit", format_rupiah(total_debit))
            with col2:
                st.metric("Total Kredit", format_rupiah(total_kredit))
            with col3:
                if total_debit == total_kredit:
                    st.success("✅ Jurnal Balance!")
                else:
                    st.error("❌ Jurnal Tidak Balance!")

            # Export option - pakai df_display
            st.download_button(
                label="📥 Export ke Excel",
                data=df_display.to_csv(index=False).encode('utf-8'),
                file_name=f"jurnal_umum_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                use_container_width=True
            )
            
            # ========== FITUR HAPUS JURNAL ==========
            st.markdown("---")
            st.markdown("### 🗑️ Hapus Transaksi Jurnal")
            
            # HAPUS PER TRANSAKSI (SEMUA AKUN DALAM SATU TRANSAKSI)
            if transaction_groups:
                col1, col2 = st.columns([2, 1])
                
                with col1:
                    # Buat opsi transaksi
                    transaction_options = {}
                    for group_id, group_data in transaction_groups.items():
                        key = f"{group_data['tanggal']} - {group_data['keterangan']} ({len(group_data['rows'])} akun)"
                        transaction_options[key] = group_id
                    
                    selected_transaction = st.selectbox(
                        "Pilih transaksi yang akan dihapus:",
                        options=list(transaction_options.keys()),
                        key="pilih_transaksi_hapus"
                    )
                
                with col2:
                    if st.button("🗑️ Hapus Transaksi", type="secondary", use_container_width=True):
                        if selected_transaction:
                            group_id = transaction_options[selected_transaction]
                            transaction_data = transaction_groups[group_id]
                            row_indices_to_delete = [entry['row_index'] for entry in transaction_data['rows']]
                            
                            if delete_journal_transaction(transaction_data['keterangan'], row_indices_to_delete):
                                st.success(f"✅ Transaksi berhasil dihapus!")
                                st.rerun()
                            else:
                                st.error("❌ Gagal menghapus transaksi!")
                        else:
                            st.error("❌ Pilih transaksi yang akan dihapus!")
                
                # Tampilkan detail transaksi terpilih
                if selected_transaction:
                    group_id = transaction_options[selected_transaction]
                    transaction_data = transaction_groups[group_id]
                    
                    st.info(f"**Detail Transaksi yang Dipilih:**")
                    st.write(f"**Tanggal:** {transaction_data['tanggal']}")
                    st.write(f"**Keterangan:** {transaction_data['keterangan']}")
                    st.write(f"**Jumlah Akun:** {len(transaction_data['rows'])}")
                    
                    st.write("**Daftar Akun:**")
                    for entry in transaction_data['rows']:
                        debit_display = format_rupiah(entry['debit']) if entry['debit'] else ''
                        kredit_display = format_rupiah(entry['kredit']) if entry['kredit'] else ''
                        
                        if debit_display and debit_display != '':
                            st.write(f"• {entry['akun']} - Debit: {debit_display}")
                        else:
                            st.write(f"• {entry['akun']} - Kredit: {kredit_display}")
            
            else:
                st.info("📝 Tidak ada transaksi untuk dihapus")
                
        else:
            st.info("📝 Belum ada transaksi jurnal")

    except Exception as e:
        st.error(f"Error: {e}")

    # Tombol untuk clear data SEMUA (fungsi lama tetap ada)
    st.markdown("---")
    st.markdown("### ⚠️ Hapus Semua Data Jurnal")
    st.warning("Tindakan ini akan menghapus SEMUA data jurnal dan tidak dapat dikembalikan!")
    
    col1, col2 = st.columns(2)
    with col1:
        if st.button("🗑️ Hapus Semua Data Jurnal", type="secondary", use_container_width=True):
            try:
                wb = Workbook()
                ws_journal = wb.active
                ws_journal.title = "Jurnal Umum"
                ws_journal.append(["Tanggal", "Akun", "Debit", "Kredit", "Keterangan"])

                ws_ledger = wb.create_sheet("Buku Besar")
                ws_ledger.append(["Akun", "Tanggal", "Keterangan", "Debit", "Kredit", "Saldo"])

                wb.save('journal_ledger.xlsx')
                st.success("Data jurnal berhasil direset!")
                st.rerun()
            except Exception as e:
                st.error(f"Error: {e}")
    
    with col2:
        if st.button("🔄 Hitung Ulang Saldo Buku Besar", type="secondary", use_container_width=True):
            if recalculate_all_ledger_balances():
                st.success("✅ Saldo Buku Besar berhasil dihitung ulang!")
            else:
                st.error("❌ Gagal menghitung ulang saldo!")

# ========== FUNGSI UNTUK HAPUS ==========
def delete_journal_transaction(keterangan, row_indices):
    """Hapus satu transaksi lengkap berdasarkan keterangan"""
    try:
        wb = load_workbook('journal_ledger.xlsx')
        ws_journal = wb['Jurnal Umum']
        ws_ledger = wb['Buku Besar']
        
        # 1. Hapus baris dari Jurnal Umum (dari belakang ke depan)
        for row_idx in sorted(row_indices, reverse=True):
            if row_idx <= ws_journal.max_row:
                ws_journal.delete_rows(row_idx)
        
        # 2. Hapus entri terkait di Buku Besar berdasarkan keterangan
        rows_to_delete_ledger = []
        
        for i, row in enumerate(ws_ledger.iter_rows(min_row=2, values_only=True), 2):
            if row and row[2] and str(row[2]) == str(keterangan):
                rows_to_delete_ledger.append(i)
        
        # Hapus dari belakang ke depan
        for row_idx in sorted(rows_to_delete_ledger, reverse=True):
            if row_idx <= ws_ledger.max_row:
                ws_ledger.delete_rows(row_idx)
        
        # 3. Hitung ulang saldo Buku Besar
        recalculate_all_ledger_balances_ws(ws_ledger)
        
        wb.save('journal_ledger.xlsx')
        wb.close()
        
        return True
        
    except Exception as e:
        st.error(f"Error dalam delete_journal_transaction: {e}")
        return False

def recalculate_all_ledger_balances_ws(ws_ledger):
    """Hitung ulang semua saldo di Buku Besar (versi dengan worksheet)"""
    try:
        # Kelompokkan data per akun
        account_transactions = {}
        
        for row in ws_ledger.iter_rows(min_row=2, values_only=False):
            if not row[0].value:  # Skip jika tidak ada akun
                continue
                
            account = row[0].value
            if account not in account_transactions:
                account_transactions[account] = []
            
            account_transactions[account].append(row)
        
        # Hitung ulang saldo untuk setiap akun
        for account, transactions in account_transactions.items():
            running_balance = 0.0
            
            for row in transactions:
                debit = safe_parse_price(row[3].value) if row[3].value else 0.0
                kredit = safe_parse_price(row[4].value) if row[4].value else 0.0
                
                # Tentukan jenis akun
                account_str = str(account)
                account_code = account_str.split(' - ')[0] if ' - ' in account_str else account_str
                
                # Akun dengan normal balance kredit: Liability (2), Equity (3), Revenue (4)
                is_credit_account = account_code.startswith(('2', '3', '4'))
                
                if is_credit_account:
                    running_balance = running_balance - debit + kredit
                else:
                    # Akun dengan normal balance debit: Asset (1), Expense (5,6)
                    running_balance = running_balance + debit - kredit
                
                # Update saldo
                row[5].value = running_balance
        
        return True
        
    except Exception as e:
        st.error(f"Error dalam recalculate_all_ledger_balances_ws: {e}")
        return False

def recalculate_all_ledger_balances():
    """Hitung ulang semua saldo di Buku Besar (versi standalone)"""
    try:
        wb = load_workbook('journal_ledger.xlsx')
        ws_ledger = wb['Buku Besar']
        
        success = recalculate_all_ledger_balances_ws(ws_ledger)
        
        wb.save('journal_ledger.xlsx')
        wb.close()
        
        return success
        
    except Exception as e:
        st.error(f"Error dalam recalculate_all_ledger_balances: {e}")
        return False

def show_buku_besar():
    st.markdown('<div class="main-header"><h1>📚 Buku Besar</h1></div>', unsafe_allow_html=True)
    
    # Pastikan file sudah ada sebelum memuat
    create_journal_workbook()
    
    try:
        wb = load_workbook('journal_ledger.xlsx')
        
        # Cek apakah sheet Buku Besar ada
        if 'Buku Besar' not in wb.sheetnames:
            # Buat sheet Buku Besar jika tidak ada
            ws_ledger = wb.create_sheet("Buku Besar")
            ws_ledger.append(["Akun", "Tanggal", "Keterangan", "Debit", "Kredit", "Saldo"])
            wb.save('journal_ledger.xlsx')
            st.info("Sheet Buku Besar berhasil dibuat")
        
        ws = wb['Buku Besar']
        
        # OTOMATIS HITUNG ULANG SEMUA SALDO
        # Kelompokkan data per akun
        account_data = {}
        for row in ws.iter_rows(min_row=2):
            account = row[0].value
            if account and account != "":
                if account not in account_data:
                    account_data[account] = []
                account_data[account].append(row)
        
        # Hitung ulang saldo untuk setiap akun dan update di Excel
        saldo_updated = False
        for account, rows in account_data.items():
            running_balance = 0.0
            for row in rows:
                debit = row[3].value or 0.0
                kredit = row[4].value or 0.0
                running_balance = running_balance + debit - kredit
                
                # Cek jika saldo perlu diupdate
                current_saldo = row[5].value or 0.0
                if abs(current_saldo - running_balance) > 0.01:  # Toleransi 0.01
                    row[5].value = running_balance
                    saldo_updated = True
        
        # Simpan jika ada perubahan
        if saldo_updated:
            wb.save('journal_ledger.xlsx')
            st.success("✅ Saldo berhasil dihitung ulang secara otomatis!")
        
        # Sekarang baca data yang sudah diupdate
        wb = load_workbook('journal_ledger.xlsx')
        ws = wb['Buku Besar']
        
        # Kumpulkan data per akun untuk ditampilkan
        ledger_entries = {}
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            
            account, date, keterangan, debit, kredit, saldo = row
            
            if account and account != "":  # Pastikan akun tidak kosong
                if account not in ledger_entries:
                    ledger_entries[account] = []
                
                ledger_entries[account].append({
                    "Tanggal": date if date else "",
                    "Keterangan": keterangan if keterangan else "",
                    "Debit": debit if debit else 0.0,
                    "Kredit": kredit if kredit else 0.0,
                    "Saldo": saldo if saldo else 0.0
                })
        
        wb.close()
        
        if ledger_entries:
            for account, entries in ledger_entries.items():
                try:
                    account_num, account_name = account.split(" - ", 1)
                except:
                    account_num = account
                    account_name = account
                
                st.markdown(f"### {account_name} ({account_num})")
                
                # Tentukan jenis akun untuk format saldo
                account_code = account_num.split()[0] if ' ' in account_num else account_num
                is_credit_account = account_code.startswith(('2', '3', '4'))  # Liability, Equity, Revenue
                
                # Format data untuk dataframe
                table_data = []
                for entry in entries:
                    # Untuk akun kredit, tampilkan saldo sebagai positif
                    display_saldo = abs(entry['Saldo']) if is_credit_account else entry['Saldo']
                    
                    table_data.append({
                        'Tanggal': entry['Tanggal'],
                        'Keterangan': entry['Keterangan'],
                        'Debit': format_rupiah(entry['Debit']) if entry['Debit'] else '',
                        'Kredit': format_rupiah(entry['Kredit']) if entry['Kredit'] else '',
                        'Saldo': format_rupiah(display_saldo)
                    })
                
                if table_data:
                    df = pd.DataFrame(table_data)
                    st.dataframe(df, use_container_width=True, hide_index=True)
                    
                    # Tampilkan saldo akhir dengan format yang benar
                    ending_balance = entries[-1]['Saldo'] if entries else 0
                    display_ending_balance = abs(ending_balance) if is_credit_account else ending_balance
                    
                    # Tentukan warna berdasarkan jenis akun
                    if is_credit_account:
                        # Untuk akun kredit, saldo normalnya kredit (positif)
                        balance_color = "#10b981"  # Hijau untuk saldo normal
                    else:
                        # Untuk akun debit, saldo normalnya debit (positif)
                        balance_color = "#10b981" if ending_balance >= 0 else "#10b981"
                    
                    # Tampilkan jenis saldo
                    saldo_type = "Kredit" if is_credit_account else "Debit"
                    
                    st.markdown(f"""
                    <div style="background: {balance_color}; color: white; padding: 0.5rem 1rem; border-radius: 5px; margin-bottom: 1rem;">
                        <strong>Saldo Akhir: {format_rupiah(display_ending_balance)} ({saldo_type})</strong>
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.info("Tidak ada transaksi untuk akun ini")
                
                st.markdown("---")
        else:
            st.info("Belum ada data buku besar. Silakan input transaksi di menu Jurnal Umum terlebih dahulu.")
    
    except Exception as e:
        st.error(f"Error: {e}")

    # Tombol untuk reset data buku besar saja
    st.markdown("---")
    if st.button("🗑️ Reset Data Buku Besar", type="secondary", use_container_width=True):
        try:
            wb = Workbook()
            ws_journal = wb.active
            ws_journal.title = "Jurnal Umum"
            ws_journal.append(["Tanggal", "Akun", "Debit", "Kredit", "Keterangan"])
            
            ws_ledger = wb.create_sheet("Buku Besar")
            ws_ledger.append(["Akun", "Tanggal", "Keterangan", "Debit", "Kredit", "Saldo"])
            
            wb.save('journal_ledger.xlsx')
            st.success("Data buku besar berhasil direset!")
            st.rerun()
        except Exception as e:
            st.error(f"Error: {e}")

def show_neraca_saldo():
    st.markdown('<div class="main-header"><h1>⚖️ Neraca Saldo</h1></div>', unsafe_allow_html=True)
    
    try:
        wb = load_workbook('journal_ledger.xlsx')
        ws = wb['Buku Besar']
        
        neraca_data = []
        
        # Kumpulkan semua akun unik dan ambil saldo terakhir
        account_balances = {}
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            
            account, date, keterangan, debit, kredit, saldo = row
            
            if account is None or account == "":
                continue
            
            # Simpan saldo terakhir untuk setiap akun
            if account not in account_balances:
                try:
                    account_num, account_name = account.split(" - ", 1)
                except:
                    account_num = account
                    account_name = account
                
                account_balances[account] = {
                    'account_num': account_num,
                    'account_name': account_name,
                    'latest_saldo': 0.0
                }
            
            # Update saldo terakhir jika ada nilai saldo
            if saldo is not None:
                account_balances[account]['latest_saldo'] = saldo
        
        wb.close()
        
        # Konversi ke format neraca saldo
        table_data = []
        total_debit = 0.0
        total_kredit = 0.0
        
        for account, data in account_balances.items():
            saldo = data['latest_saldo']
            
            # LOGIKA SEDERHANA:
            # Saldo positif → Debit, Saldo negatif → Kredit
            if saldo >= 0:
                debit_amount = saldo
                kredit_amount = 0.0
            else:
                debit_amount = 0.0
                kredit_amount = abs(saldo)
            
            total_debit += debit_amount
            total_kredit += kredit_amount
            
            table_data.append({
                'No Akun': data['account_num'],
                'Nama Akun': data['account_name'],
                'Debit': format_rupiah(debit_amount) if debit_amount > 0 else '',
                'Kredit': format_rupiah(kredit_amount) if kredit_amount > 0 else ''
            })
        
        # Urutkan berdasarkan nomor akun
        table_data.sort(key=lambda x: x['No Akun'])
        
        if table_data:
            df = pd.DataFrame(table_data)
            
            # Tampilkan dataframe dengan kolom sesuai permintaan
            st.dataframe(
                df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "No Akun": st.column_config.TextColumn("No Akun", width="small"),
                    "Nama Akun": st.column_config.TextColumn("Nama Akun", width="medium"),
                    "Debit": st.column_config.TextColumn("Debit", width="medium"),
                    "Kredit": st.column_config.TextColumn("Kredit", width="medium")
                }
            )
            
            # Total row dengan layout yang lebih baik
            st.markdown("---")
            col1, col2, col3, col4 = st.columns([2, 2, 1, 1])
            with col1:
                st.markdown("**TOTAL**")
            with col2:
                st.markdown("")  # Spacer untuk Nama Akun
            with col3:
                st.markdown(f"**{format_rupiah(total_debit)}**")
            with col4:
                st.markdown(f"**{format_rupiah(total_kredit)}**")
            
            # Balance check
            is_balanced = abs(total_debit - total_kredit) < 0.01
            st.markdown("---")
            if is_balanced:
                st.success("✅ NERACA SALDO SEIMBANG")
            else:
                st.error(f"❌ NERACA SALDO TIDAK SEIMBANG - Selisih: {format_rupiah(abs(total_debit - total_kredit))}")
        else:
            st.info("Belum ada data neraca saldo. Silakan input transaksi di menu Jurnal Umum terlebih dahulu.")
    
    except Exception as e:
        st.error(f"Error: {e}")

def show_jurnal_penyesuaian():
    st.markdown('<div class="main-header"><h1>📘 Input Jurnal Penyesuaian</h1></div>', unsafe_allow_html=True)
    create_journal_workbook() 
    
    st.markdown("### Input Jurnal Penyesuaian Baru")
    
    if 'adj_debit_accounts' not in st.session_state:
        st.session_state.adj_debit_accounts = [{'account': '', 'amount': 0}]
    if 'adj_credit_accounts' not in st.session_state:
        st.session_state.adj_credit_accounts = [{'account': '', 'amount': 0}]

    col1, col2 = st.columns(2)
    with col1:
        if st.button("➕ Tambah Akun Debit", key="adj_add_debit"):
            st.session_state.adj_debit_accounts.append({'account': '', 'amount': 0})
            st.rerun()
    with col2:
        if st.button("➕ Tambah Akun Kredit", key="adj_add_credit"):
            st.session_state.adj_credit_accounts.append({'account': '', 'amount': 0})
            st.rerun()
    
    with st.form("form_jurnal_penyesuaian", clear_on_submit=True):
        date = st.date_input("Tanggal Penyesuaian", datetime.now())
        
        st.markdown("**Akun Debit**")
        
        for i in range(len(st.session_state.adj_debit_accounts)):
            st.markdown(f"**Debit {i+1}**")
            col_acc, col_amt, col_del = st.columns([2, 1, 0.5])
            with col_acc:
                account_options = [f"{code} - {name}" for code, name in ACCOUNTS.items()]
                current_account = st.session_state.adj_debit_accounts[i]['account']
                default_index = account_options.index(current_account) if current_account in account_options else 0
                
                selected_account = st.selectbox(
                    f"Pilih Akun Debit {i+1}", 
                    account_options, 
                    key=f"adj_debit_acc_{i}",
                    index=default_index
                )
                st.session_state.adj_debit_accounts[i]['account'] = selected_account
            with col_amt:
                amount = st.number_input(
                    f"Nominal Debit {i+1}", 
                    min_value=0, 
                    value=st.session_state.adj_debit_accounts[i]['amount'],
                    key=f"adj_debit_amt_{i}"
                )
                st.session_state.adj_debit_accounts[i]['amount'] = amount
            with col_del:
                if i > 0: 
                    if st.form_submit_button("❌", type="secondary", key=f"adj_del_debit_{i}"):
                        st.session_state.adj_debit_accounts.pop(i)
                        st.rerun()
        
        st.markdown("---")
        st.markdown("**Akun Kredit**")
        
        for i in range(len(st.session_state.adj_credit_accounts)):
            st.markdown(f"**Kredit {i+1}**")
            col_acc, col_amt, col_del = st.columns([2, 1, 0.5])
            with col_acc:
                account_options = [f"{code} - {name}" for code, name in ACCOUNTS.items()]
                current_account = st.session_state.adj_credit_accounts[i]['account']
                default_index = account_options.index(current_account) if current_account in account_options else 0
                
                selected_account = st.selectbox(
                    f"Pilih Akun Kredit {i+1}", 
                    account_options, 
                    key=f"adj_credit_acc_{i}",
                    index=default_index
                )
                st.session_state.adj_credit_accounts[i]['account'] = selected_account
            with col_amt:
                amount = st.number_input(
                    f"Nominal Kredit {i+1}", 
                    min_value=0, 
                    value=st.session_state.adj_credit_accounts[i]['amount'],
                    key=f"adj_credit_amt_{i}"
                )
                st.session_state.adj_credit_accounts[i]['amount'] = amount
            with col_del:
                if i > 0:  
                    if st.form_submit_button("❌", type="secondary", key=f"adj_del_credit_{i}"):
                        st.session_state.adj_credit_accounts.pop(i)
                        st.rerun()
        
        st.markdown("---")
        keterangan = st.text_input("Keterangan Penyesuaian", placeholder="Contoh: Penyesuaian penyusutan kendaraan, Penyesuaian beban dibayar di muka, dll.")
        
        submit = st.form_submit_button("💾 Simpan Jurnal Penyesuaian", use_container_width=True)
        
        if submit:
            valid_debit_accounts = [acc for acc in st.session_state.adj_debit_accounts if acc['amount'] > 0 and acc['account']]
            valid_credit_accounts = [acc for acc in st.session_state.adj_credit_accounts if acc['amount'] > 0 and acc['account']]
            
            total_debit = sum(item['amount'] for item in valid_debit_accounts)
            total_credit = sum(item['amount'] for item in valid_credit_accounts)
            
            if total_debit == 0 or total_credit == 0:
                st.error("Nominal debit dan kredit harus lebih dari 0")
            elif total_debit != total_credit:
                st.error(f"**Tidak Balance!** Total Debit ({format_rupiah(total_debit)}) ≠ Total Kredit ({format_rupiah(total_credit)})")
            elif not valid_debit_accounts or not valid_credit_accounts:
                st.error("Harap isi minimal satu akun debit dan satu akun kredit")
            elif not keterangan:
                st.error("Keterangan penyesuaian harus diisi")
            else:
                try:
                    wb = load_workbook('journal_ledger.xlsx')
                    ws_journal = wb['Jurnal Umum']
                    ws_ledger = wb['Buku Besar']
                    
                    date_str = date.strftime('%Y-%m-%d')
                    
                    keterangan_with_label = f"[PENYESUAIAN] {keterangan}"

                    if valid_debit_accounts:
                        first_debit = valid_debit_accounts[0]
                        ws_journal.append([
                            date_str,
                            first_debit['account'],
                            first_debit['amount'], 
                            0,keterangan_with_label 
                        ])

                    for i in range(1, len(valid_debit_accounts)):
                        debit = valid_debit_accounts[i]
                        ws_journal.append([
                            "",debit['account'],
                            debit['amount'],0,"" 
                        ])

                    for credit in valid_credit_accounts:
                        ws_journal.append([
                            "",credit['account'],
                            0,credit['amount'],""
                        ])
                    
                    ws_journal.append(["", "", "", "", ""])
                    
                    current_balances = {}
                    for row in ws_ledger.iter_rows(min_row=2, values_only=True):
                        if row and row[0]: 
                            account = row[0]
                            debit = safe_parse_price(row[3]) if row[3] else 0
                            kredit = safe_parse_price(row[4]) if row[4] else 0
                            saldo = safe_parse_price(row[5]) if row[5] else 0
                            
                            if account not in current_balances:
                                current_balances[account] = saldo

                    for debit in valid_debit_accounts:
                        account = debit['account']
                        nominal = debit['amount']

                        current_balance = current_balances.get(account, 0)
                        new_balance = current_balance + nominal
                        current_balances[account] = new_balance
                        
                        ws_ledger.append([
                            account,
                            date_str,
                            keterangan_with_label,
                            nominal,0,new_balance
                        ])
                    
                    for credit in valid_credit_accounts:
                        account = credit['account']
                        nominal = credit['amount']

                        current_balance = current_balances.get(account, 0)
                        new_balance = current_balance - nominal
                        current_balances[account] = new_balance
                        
                        ws_ledger.append([
                            account,
                            date_str,
                            keterangan_with_label,
                            0,nominal,
                            new_balance
                        ])
                    
                    wb.save('journal_ledger.xlsx')
                    wb.close()
                    
                    st.success("✅ Jurnal Penyesuaian berhasil disimpan ke Jurnal Umum dan Buku Besar!")
                    
                    st.session_state.adj_debit_accounts = [{'account': '', 'amount': 0}]
                    st.session_state.adj_credit_accounts = [{'account': '', 'amount': 0}]
                    st.rerun()
                
                except Exception as e:
                    st.error(f"Error: {e}")

def show_laporan_keuangan():
    st.markdown('<div class="main-header"><h1>📋 Laporan Keuangan</h1></div>', unsafe_allow_html=True)
    
    st.markdown("### Pilih Jenis Laporan")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if st.button("📊 Laporan Laba Rugi", use_container_width=True):
            st.session_state.current_page = "Laba Rugi"
            st.rerun()
    
    with col2:
        if st.button("📈 Laporan Perubahan Modal", use_container_width=True):
            st.session_state.current_page = "Perubahan Modal"
            st.rerun()
    
    with col3:
        if st.button("💰 Laporan Posisi Keuangan", use_container_width=True):
            st.session_state.current_page = "Posisi Keuangan"
            st.rerun()

def show_laba_rugi():
    st.markdown('<div class="main-header"><h1>📊 Laporan Laba Rugi</h1></div>', unsafe_allow_html=True)
    
    if st.button("← Kembali ke Laporan Keuangan"):
        st.session_state.current_page = "Laporan Keuangan"
        st.rerun()
    
    try:
        wb = load_workbook('journal_ledger.xlsx')
        ws = wb['Buku Besar']
        
        # Kumpulkan semua akun dan saldo terakhir
        account_balances = {}
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            
            account, date, keterangan, debit, kredit, saldo = row
            
            if account is None or account == "":
                continue
            
            # Simpan saldo terakhir untuk setiap akun
            if account not in account_balances:
                try:
                    account_num, account_name = account.split(" - ", 1)
                except:
                    account_num = account
                    account_name = account
                
                account_balances[account] = {
                    'account_num': account_num.strip(),
                    'account_name': account_name,
                    'latest_saldo': 0.0
                }
            
            # Update saldo terakhir jika ada nilai saldo
            if saldo is not None:
                account_balances[account]['latest_saldo'] = saldo
        
        wb.close()
        
        # Kelompokkan akun berdasarkan kategori
        pendapatan_akun = []
        hpp_akun = []
        beban_akun = []
        
        for account, data in account_balances.items():
            account_num = data['account_num']
            saldo = data['latest_saldo']
            
            # Ambil kode utama (angka sebelum - atau spasi)
            if '-' in account_num:
                main_code = account_num.split('-')[0].strip()
            elif ' ' in account_num:
                main_code = account_num.split()[0].strip()
            else:
                main_code = account_num
            
            # Klasifikasi akun berdasarkan kode
            if main_code.startswith('4'):  # Pendapatan
                pendapatan_akun.append({
                    'nama': data['account_name'],
                    'nominal': abs(saldo)  # Untuk akun pendapatan, ambil nilai absolut
                })
            elif main_code.startswith('5'):  # HPP
                hpp_akun.append({
                    'nama': data['account_name'],
                    'nominal': abs(saldo)  # Untuk akun HPP, ambil nilai absolut
                })
            elif main_code.startswith('6'):  # Beban
                beban_akun.append({
                    'nama': data['account_name'],
                    'nominal': abs(saldo)  # Untuk akun beban, ambil nilai absolut
                })
        
        # Hitung total masing-masing kategori
        total_pendapatan = sum(akun['nominal'] for akun in pendapatan_akun)
        total_hpp = sum(akun['nominal'] for akun in hpp_akun)
        total_beban = sum(akun['nominal'] for akun in beban_akun)
        
        # Hitung laba kotor dan laba bersih
        laba_kotor = total_pendapatan - total_hpp
        laba_bersih = laba_kotor - total_beban
        
        # Tampilkan Laporan Laba Rugi
        st.markdown("### A. Pendapatan Usaha")
        
        if pendapatan_akun:
            for akun in pendapatan_akun:
                col1, col2 = st.columns([3, 1])
                with col1:
                    st.write(akun['nama'])
                with col2:
                    st.write(format_rupiah(akun['nominal']))
        else:
            col1, col2 = st.columns([3, 1])
            with col1:
                st.write("Pendapatan")
            with col2:
                st.write(format_rupiah(0))
        
        st.markdown("---")
        col1, col2 = st.columns([3, 1])
        with col1:
            st.markdown("**Total Pendapatan**")
        with col2:
            st.markdown(f"**{format_rupiah(total_pendapatan)}**")
        
        st.markdown("### B. Harga Pokok Penjualan (HPP)")
        
        if hpp_akun:
            for akun in hpp_akun:
                col1, col2 = st.columns([3, 1])
                with col1:
                    st.write(akun['nama'])
                with col2:
                    st.write(format_rupiah(akun['nominal']))
        else:
            col1, col2 = st.columns([3, 1])
            with col1:
                st.write("Harga Pokok Penjualan")
            with col2:
                st.write(format_rupiah(0))
        
        st.markdown("---")
        col1, col2 = st.columns([3, 1])
        with col1:
            st.markdown("**Total HPP**")
        with col2:
            st.markdown(f"**{format_rupiah(total_hpp)}**")
        
        st.markdown("---")
        col1, col2 = st.columns([3, 1])
        with col1:
            st.markdown("**Laba Kotor**")
        with col2:
            laba_kotor_color = "green" if laba_kotor >= 0 else "red"
            st.markdown(f"<span style='color: {laba_kotor_color}; font-weight: bold;'>{format_rupiah(laba_kotor)}</span>", unsafe_allow_html=True)
        
        st.markdown("### C. Beban Usaha")
        
        if beban_akun:
            for akun in beban_akun:
                col1, col2 = st.columns([3, 1])
                with col1:
                    st.write(akun['nama'])
                with col2:
                    st.write(format_rupiah(akun['nominal']))
        else:
            col1, col2 = st.columns([3, 1])
            with col1:
                st.write("Beban Usaha")
            with col2:
                st.write(format_rupiah(0))
        
        st.markdown("---")
        col1, col2 = st.columns([3, 1])
        with col1:
            st.markdown("**Total Beban**")
        with col2:
            st.markdown(f"**{format_rupiah(total_beban)}**")
        
        # Hasil Akhir
        st.markdown("---")
        laba_bersih_color = "green" if laba_bersih >= 0 else "red"
        st.markdown(f"""
        <div style="background: {laba_bersih_color}; color: white; padding: 1.5rem; border-radius: 10px; text-align: center; margin-top: 2rem;">
            <h2>Laba Bersih Sebelum Pajak</h2>
            <h1>{format_rupiah(laba_bersih)}</h1>
            <p>{'Perusahaan mengalami laba' if laba_bersih >= 0 else 'Perusahaan mengalami rugi'}</p>
        </div>
        """, unsafe_allow_html=True)
    
    except Exception as e:
        st.error(f"Error: {e}")

def show_perubahan_modal():
    st.markdown('<div class="main-header"><h1>📈 Laporan Perubahan Modal</h1></div>', unsafe_allow_html=True)
    
    if st.button("← Kembali ke Laporan Keuangan"):
        st.session_state.current_page = "Laporan Keuangan"
        st.rerun()
    
    try:
        wb = load_workbook('journal_ledger.xlsx')
        ws = wb['Buku Besar']
        
        # Kumpulkan semua akun dan saldo terakhir
        account_balances = {}
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            
            account, date, keterangan, debit, kredit, saldo = row
            
            if account is None or account == "":
                continue
            
            # Simpan saldo terakhir untuk setiap akun
            if account not in account_balances:
                try:
                    account_num, account_name = account.split(" - ", 1)
                except:
                    account_num = account
                    account_name = account
                
                account_balances[account] = {
                    'account_num': account_num.strip(),
                    'account_name': account_name,
                    'latest_saldo': 0.0
                }
            
            # Update saldo terakhir jika ada nilai saldo
            if saldo is not None:
                account_balances[account]['latest_saldo'] = saldo
        
        wb.close()
        
        # Ambil Modal Awal dari akun modal (3-30000)
        modal_awal = 0.0
        prive = 0.0
        
        for account, data in account_balances.items():
            account_num = data['account_num']
            saldo = data['latest_saldo']
            
            # Modal Awal (3-30000)
            if account_num == '3-30000' or account_num == '30000':
                modal_awal = abs(saldo)
            
            # Prive (3-30100)
            elif account_num == '3-30100' or account_num == '30100':
                prive = abs(saldo)
        
        # Hitung Laba Bersih dari Laporan Laba Rugi
        # Kelompokkan akun berdasarkan kategori
        pendapatan_akun = []
        hpp_akun = []
        beban_akun = []
        
        for account, data in account_balances.items():
            account_num = data['account_num']
            saldo = data['latest_saldo']
            
            # Ambil kode utama (angka sebelum - atau spasi)
            if '-' in account_num:
                main_code = account_num.split('-')[0].strip()
            elif ' ' in account_num:
                main_code = account_num.split()[0].strip()
            else:
                main_code = account_num
            
            # Klasifikasi akun berdasarkan kode
            if main_code.startswith('4'):  # Pendapatan
                pendapatan_akun.append({
                    'nama': data['account_name'],
                    'nominal': abs(saldo)
                })
            elif main_code.startswith('5'):  # HPP
                hpp_akun.append({
                    'nama': data['account_name'],
                    'nominal': abs(saldo)
                })
            elif main_code.startswith('6'):  # Beban
                beban_akun.append({
                    'nama': data['account_name'],
                    'nominal': abs(saldo)
                })
        
        # Hitung Laba Bersih
        total_pendapatan = sum(akun['nominal'] for akun in pendapatan_akun)
        total_hpp = sum(akun['nominal'] for akun in hpp_akun)
        total_beban = sum(akun['nominal'] for akun in beban_akun)
        
        laba_kotor = total_pendapatan - total_hpp
        laba_bersih = laba_kotor - total_beban
        
        # Hitung Modal Akhir
        modal_akhir = modal_awal + laba_bersih - prive
        
        # Tampilkan Laporan Perubahan Modal
        st.markdown("### Laporan Perubahan Modal")
        
        # Modal Awal
        col1, col2 = st.columns([3, 1])
        with col1:
            st.write("Modal Awal")
        with col2:
            st.write(format_rupiah(modal_awal))
        
        # Laba/Rugi Bersih
        laba_color = "green" if laba_bersih >= 0 else "red"
        laba_text = "Laba Bersih" if laba_bersih >= 0 else "Rugi Bersih"
        
        col1, col2 = st.columns([3, 1])
        with col1:
            st.markdown(f"**{laba_text}**")
        with col2:
            st.markdown(f"<span style='color: {laba_color}'>{format_rupiah(abs(laba_bersih))}</span>", unsafe_allow_html=True)
        
        # Prive
        col1, col2 = st.columns([3, 1])
        with col1:
            st.markdown("**Prive**")
        with col2:
            st.markdown(f"<span style='color: red'>({format_rupiah(prive)})</span>", unsafe_allow_html=True)
        
        st.markdown("---")
        
        # Modal Akhir
        modal_color = "green" if modal_akhir >= 0 else "red"
        st.markdown(f"""
        <div style="background: {modal_color}; color: white; padding: 1.5rem; border-radius: 10px; text-align: center; margin-top: 1rem;">
            <h2>Modal Akhir</h2>
            <h1>{format_rupiah(modal_akhir)}</h1>
        </div>
        """, unsafe_allow_html=True)
        
    except Exception as e:
        st.error(f"Error: {e}")

def show_neraca():
    st.markdown('<div class="main-header"><h1>📋 Laporan Posisi Keuangan (Neraca)</h1></div>', unsafe_allow_html=True)
    
    if st.button("← Kembali ke Laporan Keuangan"):
        st.session_state.current_page = "Laporan Keuangan"
        st.rerun()
    
    try:
        def is_accumulation_account(account_name):
            """Cek apakah akun termasuk akun akumulasi"""
            if not account_name:
                return False
            accumulation_keywords = ['akumulasi', 'accumulation', 'penyusutan', 'depreciation']
            account_lower = str(account_name).lower()
            return any(keyword in account_lower for keyword in accumulation_keywords)
        
        wb = load_workbook('journal_ledger.xlsx')
        ws = wb['Buku Besar']
        
        # Kumpulkan semua akun dan saldo terakhir
        account_balances = {}
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            
            account, date, keterangan, debit, kredit, saldo = row
            
            if account is None or account == "":
                continue
            
            if account not in account_balances:
                try:
                    account_num, account_name = account.split(" - ", 1)
                except:
                    account_num = account
                    account_name = account
                
                account_balances[account] = {
                    'account_num': account_num.strip(),
                    'account_name': account_name,
                    'latest_saldo': 0.0
                }
            
            if saldo is not None:
                account_balances[account]['latest_saldo'] = saldo
        
        wb.close()
        
        # Hitung Modal Akhir (sama seperti sebelumnya)
        modal_awal = 0.0
        prive = 0.0
        
        for account, data in account_balances.items():
            account_num = data['account_num']
            saldo = data['latest_saldo']
            
            if account_num == '3-30000' or account_num == '30000':
                modal_awal = abs(saldo)
            elif account_num == '3-30100' or account_num == '30100':
                prive = abs(saldo)
        
        # Hitung Laba Bersih
        pendapatan_akun = []
        hpp_akun = []
        beban_akun = []
        
        for account, data in account_balances.items():
            account_num = data['account_num']
            saldo = data['latest_saldo']
            
            if '-' in account_num:
                main_code = account_num.split('-')[0].strip()
            elif ' ' in account_num:
                main_code = account_num.split()[0].strip()
            else:
                main_code = account_num
            
            if main_code.startswith('4'):
                pendapatan_akun.append({
                    'nama': data['account_name'],
                    'nominal': abs(saldo)
                })
            elif main_code.startswith('5'):
                hpp_akun.append({
                    'nama': data['account_name'],
                    'nominal': abs(saldo)
                })
            elif main_code.startswith('6'):
                beban_akun.append({
                    'nama': data['account_name'],
                    'nominal': abs(saldo)
                })
        
        total_pendapatan = sum(akun['nominal'] for akun in pendapatan_akun)
        total_hpp = sum(akun['nominal'] for akun in hpp_akun)
        total_beban = sum(akun['nominal'] for akun in beban_akun)
        
        laba_kotor = total_pendapatan - total_hpp
        laba_bersih = laba_kotor - total_beban
        
        modal_akhir = modal_awal + laba_bersih - prive
        
        # PERBAIKAN: Kelompokkan akun dengan handling akumulasi penyusutan yang benar
        aset_lancar = []
        aset_tetap_bruto = []      
        akumulasi_penyusutan = []  
        liabilitas_pendek = []
        liabilitas_panjang = []
        
        for account, data in account_balances.items():
            account_num = data['account_num']
            saldo = data['latest_saldo']
            nominal = abs(saldo)
            
            # Klasifikasi berdasarkan kode akun
            if account_num.startswith('1-1') or account_num.startswith('11'):  # Aset Lancar
                aset_lancar.append({
                    'nama': data['account_name'],
                    'nominal': nominal
                })
            elif account_num.startswith('1-2') or account_num.startswith('12'):  # Aset Tetap Bruto
                # Pisahkan antara aset tetap dan akumulasi penyusutan
                if is_accumulation_account(data['account_name']):
                    akumulasi_penyusutan.append({
                        'nama': data['account_name'],
                        'nominal': nominal
                    })
                else:
                    aset_tetap_bruto.append({
                        'nama': data['account_name'],
                        'nominal': nominal
                    })
            elif account_num.startswith('2-1') or account_num.startswith('21'):  # Liabilitas Jangka Pendek
                liabilitas_pendek.append({
                    'nama': data['account_name'],
                    'nominal': nominal
                })
            elif account_num.startswith('2-2') or account_num.startswith('22'):  # Liabilitas Jangka Panjang
                liabilitas_panjang.append({
                    'nama': data['account_name'],
                    'nominal': nominal
                })
        
        # PERBAIKAN: Hitung aset tetap NETO (Bruto - Akumulasi Penyusutan)
        subtotal_aset_lancar = sum(akun['nominal'] for akun in aset_lancar)
        subtotal_aset_tetap_bruto = sum(akun['nominal'] for akun in aset_tetap_bruto)
        total_akumulasi_penyusutan = sum(akun['nominal'] for akun in akumulasi_penyusutan)
        subtotal_aset_tetap_neto = subtotal_aset_tetap_bruto - total_akumulasi_penyusutan
        
        # Total Aset = Aset Lancar + Aset Tetap Neto
        total_aset = subtotal_aset_lancar + subtotal_aset_tetap_neto
        
        subtotal_liabilitas_pendek = sum(akun['nominal'] for akun in liabilitas_pendek)
        subtotal_liabilitas_panjang = sum(akun['nominal'] for akun in liabilitas_panjang)
        total_liabilitas = subtotal_liabilitas_pendek + subtotal_liabilitas_panjang
        
        total_ekuitas = modal_akhir
        total_liabilitas_ekuitas = total_liabilitas + total_ekuitas
        
        # Tampilkan Laporan Posisi Keuangan dengan format yang benar
        st.markdown("""
        <style>
        .neraca-container {
            display: flex;
            justify-content: space-between;
            gap: 2rem;
        }
        .neraca-column {
            flex: 1;
            background: white;
            padding: 1.5rem;
            border-radius: 10px;
            border: 1px solid #e2e8f0;
        }
        .neraca-header {
            text-align: center;
            font-weight: bold;
            font-size: 1.2rem;
            margin-bottom: 1.5rem;
            padding-bottom: 0.5rem;
            border-bottom: 2px solid #DDB27A;
        }
        .neraca-item {
            display: flex;
            justify-content: space-between;
            padding: 0.5rem 0;
            border-bottom: 1px solid #f1f5f9;
        }
        .neraca-subtotal {
            display: flex;
            justify-content: space-between;
            font-weight: bold;
            padding: 0.75rem 0;
            border-top: 2px solid #e2e8f0;
            border-bottom: 2px solid #e2e8f0;
            margin: 0.5rem 0;
        }
        .neraca-total {
            display: flex;
            justify-content: space-between;
            font-weight: bold;
            font-size: 1.1rem;
            padding: 1rem 0;
            border-top: 3px double #DDB27A;
            border-bottom: 3px double #DDB27A;
            margin: 1rem 0;
            background: #FDF3B9;
        }
        .neraca-group {
            margin-bottom: 1.5rem;
        }
        .group-title {
            font-weight: bold;
            color: #475569;
            margin-bottom: 0.5rem;
            padding-left: 0.5rem;
            border-left: 3px solid #DDB27A;
        }
        .akumulasi-item {
            display: flex;
            justify-content: space-between;
            padding: 0.5rem 0;
            border-bottom: 1px solid #f1f5f9;
            color: #666;
            font-style: italic;
            padding-left: 1rem;
        }
        </style>
        """, unsafe_allow_html=True)
        
        # Container utama untuk dua kolom
        st.markdown('<div class="neraca-container">', unsafe_allow_html=True)
        
        # KOLOM KIRI - ASET
        st.markdown('<div class="neraca-column">', unsafe_allow_html=True)
        st.markdown('<div class="neraca-header">ASET</div>', unsafe_allow_html=True)
        
        # Aset Lancar
        st.markdown('<div class="neraca-group">', unsafe_allow_html=True)
        st.markdown('<div class="group-title">Aset Lancar</div>', unsafe_allow_html=True)
        
        if aset_lancar:
            for akun in aset_lancar:
                st.markdown(f"""
                <div class="neraca-item">
                    <span>{akun['nama']}</span>
                    <span>{format_rupiah(akun['nominal'])}</span>
                </div>
                """, unsafe_allow_html=True)
        else:
            st.markdown("""
            <div class="neraca-item">
                <span>Tidak ada aset lancar</span>
                <span>-</span>
            </div>
            """, unsafe_allow_html=True)
        
        # Subtotal Aset Lancar
        st.markdown(f"""
        <div class="neraca-subtotal">
            <span>Subtotal Aset Lancar</span>
            <span>{format_rupiah(subtotal_aset_lancar)}</span>
        </div>
        """, unsafe_allow_html=True)
        
        # Aset Tetap - PERBAIKAN: Tampilkan Bruto, Akumulasi, dan Neto
        st.markdown('<div class="neraca-group">', unsafe_allow_html=True)
        st.markdown('<div class="group-title">Aset Tetap</div>', unsafe_allow_html=True)
        
        if aset_tetap_bruto:
            for akun in aset_tetap_bruto:
                st.markdown(f"""
                <div class="neraca-item">
                    <span>{akun['nama']}</span>
                    <span>{format_rupiah(akun['nominal'])}</span>
                </div>
                """, unsafe_allow_html=True)
        else:
            st.markdown("""
            <div class="neraca-item">
                <span>Tidak ada aset tetap</span>
                <span>-</span>
            </div>
            """, unsafe_allow_html=True)
        
        # Tampilkan Akumulasi Penyusutan sebagai pengurang
        if akumulasi_penyusutan:
            st.markdown('<div style="margin-top: 0.5rem;">', unsafe_allow_html=True)
            for akun in akumulasi_penyusutan:
                st.markdown(f"""
                <div class="akumulasi-item">
                    <span>{akun['nama']}</span>
                    <span>({format_rupiah(akun['nominal'])})</span>
                </div>
                """, unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)
        
        # Subtotal Aset Tetap Bruto
        st.markdown(f"""
        <div class="neraca-subtotal">
            <span>Aset Tetap Bruto</span>
            <span>{format_rupiah(subtotal_aset_tetap_bruto)}</span>
        </div>
        """, unsafe_allow_html=True)
        
        # Akumulasi Penyusutan
        if akumulasi_penyusutan:
            st.markdown(f"""
            <div class="neraca-item">
                <span>Akumulasi Penyusutan</span>
                <span>({format_rupiah(total_akumulasi_penyusutan)})</span>
            </div>
            """, unsafe_allow_html=True)
        
        # Subtotal Aset Tetap Neto
        st.markdown(f"""
        <div class="neraca-subtotal">
            <span>Aset Tetap Neto</span>
            <span>{format_rupiah(subtotal_aset_tetap_neto)}</span>
        </div>
        """, unsafe_allow_html=True)
        
        # Total Aset
        st.markdown(f"""
        <div class="neraca-total">
            <span>TOTAL ASET</span>
            <span>{format_rupiah(total_aset)}</span>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown('</div>', unsafe_allow_html=True)  # Tutup kolom aset
        
        # KOLOM KANAN - LIABILITAS & EKUITAS
        st.markdown('<div class="neraca-column">', unsafe_allow_html=True)
        st.markdown('<div class="neraca-header">LIABILITAS & EKUITAS</div>', unsafe_allow_html=True)
        
        # Liabilitas Jangka Pendek
        st.markdown('<div class="neraca-group">', unsafe_allow_html=True)
        st.markdown('<div class="group-title">Liabilitas Jangka Pendek</div>', unsafe_allow_html=True)
        
        if liabilitas_pendek:
            for akun in liabilitas_pendek:
                st.markdown(f"""
                <div class="neraca-item">
                    <span>{akun['nama']}</span>
                    <span>{format_rupiah(akun['nominal'])}</span>
                </div>
                """, unsafe_allow_html=True)
        else:
            st.markdown("""
            <div class="neraca-item">
                <span>Tidak ada liabilitas jangka pendek</span>
                <span>-</span>
            </div>
            """, unsafe_allow_html=True)
        
        # Subtotal Liabilitas Jangka Pendek
        st.markdown(f"""
        <div class="neraca-subtotal">
            <span>Subtotal Liabilitas Jangka Pendek</span>
            <span>{format_rupiah(subtotal_liabilitas_pendek)}</span>
        </div>
        """, unsafe_allow_html=True)
        
        # Liabilitas Jangka Panjang
        st.markdown('<div class="neraca-group">', unsafe_allow_html=True)
        st.markdown('<div class="group-title">Liabilitas Jangka Panjang</div>', unsafe_allow_html=True)
        
        if liabilitas_panjang:
            for akun in liabilitas_panjang:
                st.markdown(f"""
                <div class="neraca-item">
                    <span>{akun['nama']}</span>
                    <span>{format_rupiah(akun['nominal'])}</span>
                </div>
                """, unsafe_allow_html=True)
        else:
            st.markdown("""
            <div class="neraca-item">
                <span>Tidak ada liabilitas jangka panjang</span>
                <span>-</span>
            </div>
            """, unsafe_allow_html=True)
        
        # Subtotal Liabilitas Jangka Panjang
        st.markdown(f"""
        <div class="neraca-subtotal">
            <span>Subtotal Liabilitas Jangka Panjang</span>
            <span>{format_rupiah(subtotal_liabilitas_panjang)}</span>
        </div>
        """, unsafe_allow_html=True)
        
        # Total Liabilitas
        st.markdown(f"""
        <div class="neraca-subtotal">
            <span>Total Liabilitas</span>
            <span>{format_rupiah(total_liabilitas)}</span>
        </div>
        """, unsafe_allow_html=True)
        
        # Ekuitas
        st.markdown('<div class="neraca-group">', unsafe_allow_html=True)
        st.markdown('<div class="group-title">Ekuitas</div>', unsafe_allow_html=True)
        
        st.markdown(f"""
        <div class="neraca-item">
            <span>Modal Akhir</span>
            <span>{format_rupiah(modal_akhir)}</span>
        </div>
        """, unsafe_allow_html=True)
        
        # Subtotal Ekuitas
        st.markdown(f"""
        <div class="neraca-subtotal">
            <span>Total Ekuitas</span>
            <span>{format_rupiah(total_ekuitas)}</span>
        </div>
        """, unsafe_allow_html=True)
        
        # Total Liabilitas & Ekuitas
        st.markdown(f"""
        <div class="neraca-total">
            <span>TOTAL LIABILITAS & EKUITAS</span>
            <span>{format_rupiah(total_liabilitas_ekuitas)}</span>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown('</div>', unsafe_allow_html=True)  # Tutup kolom liabilitas & ekuitas
        st.markdown('</div>', unsafe_allow_html=True)  # Tutup container utama
        
        # Balance Check
        is_balanced = abs(total_aset - total_liabilitas_ekuitas) < 0.01
        st.markdown("---")
        if is_balanced:
            st.success("✅ NERACA SEIMBANG - Total Aset = Total Liabilitas & Ekuitas")
        else:
            st.error(f"❌ NERACA TIDAK SEIMBANG - Selisih: {format_rupiah(abs(total_aset - total_liabilitas_ekuitas))}")
            
            # Debug info untuk membantu troubleshooting
            st.warning("🔍 **Debug Info:**")
            col1, col2 = st.columns(2)
            with col1:
                st.write(f"Total Aset: {format_rupiah(total_aset)}")
                st.write(f"- Aset Lancar: {format_rupiah(subtotal_aset_lancar)}")
                st.write(f"- Aset Tetap Bruto: {format_rupiah(subtotal_aset_tetap_bruto)}")
                st.write(f"- Akumulasi Penyusutan: {format_rupiah(total_akumulasi_penyusutan)}")
                st.write(f"- Aset Tetap Neto: {format_rupiah(subtotal_aset_tetap_neto)}")
            with col2:
                st.write(f"Total Liabilitas & Ekuitas: {format_rupiah(total_liabilitas_ekuitas)}")
                st.write(f"- Liabilitas: {format_rupiah(total_liabilitas)}")
                st.write(f"- Ekuitas: {format_rupiah(total_ekuitas)}")
    
    except Exception as e:
        st.error(f"Error: {e}")

if st.session_state.current_page == "Dashboard":
    show_dashboard()
elif st.session_state.current_page == "Kartu Persediaan":
    show_kartu_persediaan()
elif st.session_state.current_page == "Ringkasan Penjualan":
    show_ringkasan_penjualan()
elif st.session_state.current_page == "Jurnal Umum":
    show_jurnal_umum()
elif st.session_state.current_page == "View Jurnal":
    show_view_jurnal()
elif st.session_state.current_page == "Jurnal Penyesuaian":
    show_jurnal_penyesuaian()
elif st.session_state.current_page == "Buku Besar":
    show_buku_besar()
elif st.session_state.current_page == "Neraca Saldo":
    show_neraca_saldo()
elif st.session_state.current_page == "Laporan Keuangan":
    show_laporan_keuangan()
elif st.session_state.current_page == "Laba Rugi":
    show_laba_rugi()
elif st.session_state.current_page == "Perubahan Modal":
    show_perubahan_modal()
elif st.session_state.current_page == "Posisi Keuangan":
    show_neraca()
